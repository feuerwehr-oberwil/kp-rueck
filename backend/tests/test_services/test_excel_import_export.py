"""Tests for the excel_import_export service module.

Tests Excel template generation, validation, parsing, import, and export including:
- Template generation with example data
- Excel file validation and parsing
- Data import with different modes (replace, merge, append)
- Data export to Excel
"""

import io
from uuid import uuid4

import pytest
import pytest_asyncio
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Material, Personnel, User, Vehicle
from app.services.excel_import_export import (
    EXAMPLE_ROW_MARKER,
    MATERIAL_COLUMNS,
    PERSONNEL_COLUMNS,
    PERSONNEL_STATUSES,
    VEHICLE_COLUMNS,
    VEHICLE_STATUSES,
    ExcelImportError,
    ParsedImport,
    ParsedSheet,
    export_data_to_excel,
    generate_empty_template,
    import_data,
    validate_and_parse_excel,
)

# ============================================
# Fixtures
# ============================================


def full_workbook(
    *,
    personnel: list[dict] | None = None,
    vehicles: list[dict] | None = None,
    materials: list[dict] | None = None,
) -> ParsedImport:
    """What the parser returns for a file that has all three sheets in it.

    Empty here means "the sheet is there with only its header" – which in `replace`
    mode legitimately clears that table. A sheet the file does not have at all is a
    different thing entirely (`ParsedSheet(present=False)`) and is built by hand in
    the tests that care.
    """
    return ParsedImport(
        personnel=ParsedSheet(present=True, rows=personnel or []),
        vehicles=ParsedSheet(present=True, rows=vehicles or []),
        materials=ParsedSheet(present=True, rows=materials or []),
    )


@pytest_asyncio.fixture
async def excel_user(db_session: AsyncSession) -> User:
    """Create a test user for excel import/export tests."""
    user = User(
        id=uuid4(),
        username="excel_test_user",
        password_hash="$2b$12$test",
        role="editor",
    )
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    return user


@pytest_asyncio.fixture
async def sample_personnel(db_session: AsyncSession) -> list[Personnel]:
    """Create sample personnel for export tests."""
    # Note: Database constraint uses 'unavailable' not 'not-available'
    personnel_list = [
        Personnel(id=uuid4(), name="Alice Test", role="Fahrer", status="available"),
        Personnel(id=uuid4(), name="Bob Test", role="Atemschutz", status="unavailable"),
        Personnel(id=uuid4(), name="Charlie Test", status="available"),
    ]
    for p in personnel_list:
        db_session.add(p)
    await db_session.commit()
    for p in personnel_list:
        await db_session.refresh(p)
    return personnel_list


@pytest_asyncio.fixture
async def sample_vehicles(db_session: AsyncSession) -> list[Vehicle]:
    """Create sample vehicles for export tests."""
    vehicles_list = [
        Vehicle(
            id=uuid4(),
            name="Test TLF",
            type="TLF",
            display_order=1,
            status="available",
            radio_call_sign="Test 1",
        ),
        Vehicle(
            id=uuid4(),
            name="Test DLK",
            type="DLK",
            display_order=2,
            status="unavailable",
            radio_call_sign="Test 2",
        ),
    ]
    for v in vehicles_list:
        db_session.add(v)
    await db_session.commit()
    for v in vehicles_list:
        await db_session.refresh(v)
    return vehicles_list


@pytest_asyncio.fixture
async def sample_materials(db_session: AsyncSession) -> list[Material]:
    """Create sample materials for export tests."""
    materials_list = [
        Material(id=uuid4(), name="Test Pump 1", type="Tauchpumpen", location="TLF"),
        Material(id=uuid4(), name="Test Pump 2", type="Tauchpumpen", location="TLF"),
        Material(id=uuid4(), name="Test Tool", type="Werkzeug", location="RW", description="A test tool"),
    ]
    for m in materials_list:
        db_session.add(m)
    await db_session.commit()
    for m in materials_list:
        await db_session.refresh(m)
    return materials_list


def create_valid_excel_bytes(
    personnel: list[dict] | None = None,
    vehicles: list[dict] | None = None,
    materials: list[dict] | None = None,
) -> bytes:
    """Helper to create valid Excel file bytes for testing."""
    wb = Workbook()
    wb.remove(wb.active)

    # Personnel sheet
    ws_personnel = wb.create_sheet("Personnel")
    ws_personnel.append([col[0] for col in PERSONNEL_COLUMNS])
    for row in personnel or []:
        ws_personnel.append([row.get("name"), row.get("role"), row.get("status")])

    # Vehicles sheet
    ws_vehicles = wb.create_sheet("Vehicles")
    ws_vehicles.append([col[0] for col in VEHICLE_COLUMNS])
    for row in vehicles or []:
        ws_vehicles.append(
            [
                row.get("name"),
                row.get("type"),
                row.get("display_order"),
                row.get("status"),
                row.get("radio_call_sign"),
            ]
        )

    # Materials sheet
    ws_materials = wb.create_sheet("Materials")
    ws_materials.append([col[0] for col in MATERIAL_COLUMNS])
    for row in materials or []:
        ws_materials.append([row.get("name"), row.get("type"), row.get("location"), row.get("description")])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ============================================
# generate_empty_template Tests
# ============================================


class TestGenerateEmptyTemplate:
    """Tests for generate_empty_template function."""

    def test_returns_bytesio(self):
        """Test returns a BytesIO buffer."""
        result = generate_empty_template()
        assert isinstance(result, io.BytesIO)

    def test_creates_valid_workbook(self):
        """Test creates a valid Excel workbook."""
        result = generate_empty_template()
        wb = load_workbook(result)
        assert wb is not None

    def test_contains_all_sheets(self):
        """Test contains Personnel, Vehicles, and Materials sheets."""
        result = generate_empty_template()
        wb = load_workbook(result)
        assert "Personnel" in wb.sheetnames
        assert "Vehicles" in wb.sheetnames
        assert "Materials" in wb.sheetnames

    def test_personnel_sheet_has_correct_headers(self):
        """Test Personnel sheet has correct column headers."""
        result = generate_empty_template()
        wb = load_workbook(result)
        ws = wb["Personnel"]
        headers = [cell.value for cell in ws[1]]
        expected = [col[0] for col in PERSONNEL_COLUMNS]
        assert headers == expected

    def test_vehicles_sheet_has_correct_headers(self):
        """Test Vehicles sheet has correct column headers."""
        result = generate_empty_template()
        wb = load_workbook(result)
        ws = wb["Vehicles"]
        headers = [cell.value for cell in ws[1]]
        expected = [col[0] for col in VEHICLE_COLUMNS]
        assert headers == expected

    def test_materials_sheet_has_correct_headers(self):
        """Test Materials sheet has correct column headers."""
        result = generate_empty_template()
        wb = load_workbook(result)
        ws = wb["Materials"]
        headers = [cell.value for cell in ws[1]]
        expected = [col[0] for col in MATERIAL_COLUMNS]
        assert headers == expected

    def test_contains_example_data(self):
        """Test sheets contain example data rows."""
        result = generate_empty_template()
        wb = load_workbook(result)

        # Personnel should have 2 example rows + header
        ws_personnel = wb["Personnel"]
        assert ws_personnel.max_row >= 3

        # Vehicles should have 2 example rows + header
        ws_vehicles = wb["Vehicles"]
        assert ws_vehicles.max_row >= 3

        # Materials should have 3 example rows + header
        ws_materials = wb["Materials"]
        assert ws_materials.max_row >= 4

    def test_every_example_row_is_marked(self):
        """No example row may be readable as data by the operator scanning the file.

        A tester asked outright whether `Max Mustermann` would be imported and could
        not tell from the sheet. Every example row now says what it is, in the first
        column, in the operator's language.
        """
        wb = load_workbook(generate_empty_template())

        for sheet in ("Personnel", "Vehicles", "Materials"):
            ws = wb[sheet]
            names = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
            assert names, f"{sheet} has no example rows"
            assert all(str(name).startswith(EXAMPLE_ROW_MARKER) for name in names), names

    def test_buffer_position_at_start(self):
        """Test returned buffer is positioned at start."""
        result = generate_empty_template()
        assert result.tell() == 0


# ============================================
# validate_and_parse_excel Tests
# ============================================


class TestValidateAndParseExcel:
    """Tests for validate_and_parse_excel function."""

    def test_raises_error_for_invalid_file(self):
        """Test raises ExcelImportError for invalid file bytes."""
        with pytest.raises(ExcelImportError, match="nicht als Excel-Arbeitsmappe"):
            validate_and_parse_excel(b"not a valid excel file")

    def test_parses_empty_sheets(self):
        """Test handles empty sheets (headers only)."""
        file_bytes = create_valid_excel_bytes()
        result = validate_and_parse_excel(file_bytes)
        assert result.personnel.rows == []
        assert result.vehicles.rows == []
        assert result.materials.rows == []

    def test_parses_personnel_data(self):
        """Test parses Personnel data correctly."""
        file_bytes = create_valid_excel_bytes(
            personnel=[
                {"name": "Test Person", "role": "Fahrer", "status": "available"},
            ]
        )
        result = validate_and_parse_excel(file_bytes)
        assert len(result.personnel.rows) == 1
        assert result.personnel.rows[0]["name"] == "Test Person"
        assert result.personnel.rows[0]["role"] == "Fahrer"
        assert result.personnel.rows[0]["status"] == "available"

    def test_parses_vehicles_data(self):
        """Test parses Vehicles data correctly."""
        file_bytes = create_valid_excel_bytes(
            vehicles=[
                {
                    "name": "Test Vehicle",
                    "type": "TLF",
                    "display_order": 1,
                    "status": "available",
                    "radio_call_sign": "Test 1",
                },
            ]
        )
        result = validate_and_parse_excel(file_bytes)
        assert len(result.vehicles.rows) == 1
        assert result.vehicles.rows[0]["name"] == "Test Vehicle"
        assert result.vehicles.rows[0]["type"] == "TLF"
        assert result.vehicles.rows[0]["display_order"] == 1
        assert result.vehicles.rows[0]["status"] == "available"

    def test_parses_materials_data(self):
        """Test parses Materials data correctly."""
        file_bytes = create_valid_excel_bytes(
            materials=[
                {"name": "Test Material", "type": "Pumps", "location": "TLF", "description": "A test"},
            ]
        )
        result = validate_and_parse_excel(file_bytes)
        assert len(result.materials.rows) == 1
        assert result.materials.rows[0]["name"] == "Test Material"
        assert result.materials.rows[0]["type"] == "Pumps"
        assert result.materials.rows[0]["location"] == "TLF"

    def test_skips_empty_rows(self):
        """Test skips completely empty rows."""
        # Create file with empty row between data
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Personnel")
        ws.append(["name", "role", "status"])
        ws.append(["Person 1", "Role", "available"])
        ws.append([None, None, None])  # Empty row
        ws.append(["Person 2", "Role", "available"])

        ws_v = wb.create_sheet("Vehicles")
        ws_v.append(["name", "type", "display_order", "status", "radio_call_sign"])

        ws_m = wb.create_sheet("Materials")
        ws_m.append(["name", "type", "location", "description"])

        buffer = io.BytesIO()
        wb.save(buffer)
        file_bytes = buffer.getvalue()

        result = validate_and_parse_excel(file_bytes)
        assert len(result.personnel.rows) == 2

    def test_accepts_legacy_availability_header(self):
        """A workbook exported before the rename still imports.

        The personnel column was called "availability" until the field was renamed to
        `status`; a station with an older export sitting in its Downloads folder must not
        get "Expected columns" thrown at it.
        """
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Personnel")
        ws.append(["name", "role", "availability"])
        ws.append(["Alte Datei", "Mannschaft", "available"])

        wb.create_sheet("Vehicles").append(["name", "type", "display_order", "status", "radio_call_sign"])
        wb.create_sheet("Materials").append(["name", "type", "location", "description"])

        buffer = io.BytesIO()
        wb.save(buffer)

        result = validate_and_parse_excel(buffer.getvalue())
        assert result.personnel.rows == [{"name": "Alte Datei", "role": "Mannschaft", "status": "available"}]

    def test_validates_personnel_name_required(self):
        """Test validates Personnel name is required."""
        file_bytes = create_valid_excel_bytes(personnel=[{"name": None, "role": "Test", "status": "available"}])
        with pytest.raises(ExcelImportError, match="Spalte 'name' ist leer"):
            validate_and_parse_excel(file_bytes)

    def test_validates_personnel_status_enum(self):
        """Test validates Personnel status is valid enum."""
        file_bytes = create_valid_excel_bytes(personnel=[{"name": "Test", "role": "Test", "status": "invalid_status"}])
        with pytest.raises(ExcelImportError, match="ungültiger Status"):
            validate_and_parse_excel(file_bytes)

    def test_sets_default_personnel_status(self):
        """Test sets default status to 'unavailable' when empty."""
        file_bytes = create_valid_excel_bytes(personnel=[{"name": "Test Person", "role": "Test", "status": None}])
        result = validate_and_parse_excel(file_bytes)
        assert result.personnel.rows[0]["status"] == "unavailable"

    def test_validates_vehicle_required_fields(self):
        """Test validates all Vehicle required fields."""
        required_fields = ["name", "type", "display_order", "status", "radio_call_sign"]
        for field in required_fields:
            vehicle = {
                "name": "Test",
                "type": "TLF",
                "display_order": 1,
                "status": "available",
                "radio_call_sign": "Test 1",
            }
            vehicle[field] = None  # Remove required field
            file_bytes = create_valid_excel_bytes(vehicles=[vehicle])
            with pytest.raises(ExcelImportError, match=f"Spalte '{field}' ist leer"):
                validate_and_parse_excel(file_bytes)

    def test_validates_vehicle_status_enum(self):
        """Test validates Vehicle status is valid enum."""
        file_bytes = create_valid_excel_bytes(
            vehicles=[
                {
                    "name": "Test",
                    "type": "TLF",
                    "display_order": 1,
                    "status": "invalid_status",
                    "radio_call_sign": "Test 1",
                }
            ]
        )
        with pytest.raises(ExcelImportError, match="ungültiger Status"):
            validate_and_parse_excel(file_bytes)

    def test_validates_vehicle_display_order_is_integer(self):
        """Test validates Vehicle display_order is integer."""
        file_bytes = create_valid_excel_bytes(
            vehicles=[
                {
                    "name": "Test",
                    "type": "TLF",
                    "display_order": "not_a_number",
                    "status": "available",
                    "radio_call_sign": "Test 1",
                }
            ]
        )
        with pytest.raises(ExcelImportError, match="muss eine ganze Zahl sein"):
            validate_and_parse_excel(file_bytes)

    def test_validates_material_required_fields(self):
        """Test validates all Material required fields."""
        required_fields = ["name", "type", "location"]
        for field in required_fields:
            material = {"name": "Test", "type": "Pumps", "location": "TLF", "description": "Test"}
            material[field] = None  # Remove required field
            file_bytes = create_valid_excel_bytes(materials=[material])
            with pytest.raises(ExcelImportError, match=f"Spalte '{field}' ist leer"):
                validate_and_parse_excel(file_bytes)

    def test_validates_incorrect_column_headers(self):
        """Test raises error for incorrect column headers."""
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Personnel")
        ws.append(["wrong_column", "bad_header", "invalid"])

        ws_v = wb.create_sheet("Vehicles")
        ws_v.append(["name", "type", "display_order", "status", "radio_call_sign"])

        ws_m = wb.create_sheet("Materials")
        ws_m.append(["name", "type", "location", "description"])

        buffer = io.BytesIO()
        wb.save(buffer)

        with pytest.raises(ExcelImportError, match="falsche Spaltenüberschriften"):
            validate_and_parse_excel(buffer.getvalue())

    def test_accepts_any_material_type(self):
        """Test accepts any non-empty string for material type."""
        file_bytes = create_valid_excel_bytes(
            materials=[{"name": "Test", "type": "Custom Type 123", "location": "TLF", "description": ""}]
        )
        result = validate_and_parse_excel(file_bytes)
        assert result.materials.rows[0]["type"] == "Custom Type 123"

    def test_accepts_any_vehicle_type(self):
        """Vehicle `type` is free text, and stays free text.

        A `VEHICLE_TYPES` list used to sit next to the status enums looking like an
        allowed-values check while enforcing nothing – so the docs had to explain that
        `type` is not validated. The list is gone; this is the behaviour it never had.
        """
        file_bytes = create_valid_excel_bytes(
            vehicles=[
                {
                    "name": "Anhänger Pumpe",
                    "type": "Anhänger",
                    "display_order": 1,
                    "status": "available",
                    "radio_call_sign": "Florian 9",
                }
            ]
        )
        result = validate_and_parse_excel(file_bytes)
        assert result.vehicles.rows[0]["type"] == "Anhänger"

    def test_absent_sheet_is_not_an_empty_sheet(self):
        """A missing sheet and a header-only sheet are opposite instructions.

        Both used to arrive here as `[]`, which is how a Personnel-only workbook in
        `replace` mode deleted a station's whole fleet and material inventory.
        """
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Personnel").append(["name", "role", "status"])
        wb.create_sheet("Vehicles").append(["name", "type", "display_order", "status", "radio_call_sign"])

        buffer = io.BytesIO()
        wb.save(buffer)
        result = validate_and_parse_excel(buffer.getvalue())

        assert result.personnel.present is True
        assert result.vehicles == ParsedSheet(present=True, rows=[])
        assert result.materials == ParsedSheet(present=False, rows=[])

    def test_example_rows_are_skipped_and_real_rows_are_not(self):
        """The template's examples never reach the database, whoever left them there.

        This is the half of the fix that removes the harm: an operator who types their
        roster underneath the examples and uploads it must not end up with two fictional
        firefighters, two fictional vehicles and three fictional pumps on the board.
        """
        file_bytes = create_valid_excel_bytes(
            personnel=[
                {"name": f"{EXAMPLE_ROW_MARKER}: Max Mustermann", "role": "Fahrer", "status": "available"},
                {"name": "Beatrice Roth", "role": "", "status": "available"},
            ],
            vehicles=[
                {
                    "name": f"{EXAMPLE_ROW_MARKER}: TLF 1",
                    "type": "TLF",
                    "display_order": 1,
                    "status": "available",
                    "radio_call_sign": "Florian 1",
                },
                {
                    "name": "MTW Oberwil",
                    "type": "MTW",
                    "display_order": 2,
                    "status": "available",
                    "radio_call_sign": "Florian 9",
                },
            ],
            materials=[
                {"name": f"{EXAMPLE_ROW_MARKER}: Tauchpumpe Gr.", "type": "Tauchpumpen", "location": "TLF"},
                {"name": "Motorsäge", "type": "Werkzeug", "location": "Pio"},
            ],
        )
        result = validate_and_parse_excel(file_bytes)

        assert [row["name"] for row in result.personnel.rows] == ["Beatrice Roth"]
        assert [row["name"] for row in result.vehicles.rows] == ["MTW Oberwil"]
        assert [row["name"] for row in result.materials.rows] == ["Motorsäge"]

    def test_example_marker_is_matched_forgivingly(self):
        """The row travels through the operator's spreadsheet before it comes back.

        Case and stray whitespace are exactly what changes on the way, and a marker
        that only matched byte-for-byte would let the example through in the files
        where somebody had been editing around it.
        """
        file_bytes = create_valid_excel_bytes(
            personnel=[
                {"name": f"  {EXAMPLE_ROW_MARKER}: Max Mustermann  ", "role": "", "status": "available"},
                {"name": f"{EXAMPLE_ROW_MARKER.upper()}: Anna Schmidt", "role": "", "status": "available"},
                {"name": f"{EXAMPLE_ROW_MARKER.lower()}: Anna Schmidt", "role": "", "status": "available"},
            ],
        )
        assert validate_and_parse_excel(file_bytes).personnel == ParsedSheet(present=True, rows=[])

    def test_sheet_of_only_example_rows_is_present_not_absent(self):
        """Untouched examples mean "nothing here", not "no such sheet".

        The two are opposite instructions in `replace` mode: a present-but-empty sheet
        clears its table, while an absent one is refused with a 409 (`_refuse_missing_sheets`).
        An operator who deleted nothing at all would have no way to act on that refusal,
        so the skip must never turn one case into the other.
        """
        result = validate_and_parse_excel(generate_empty_template().getvalue())

        assert result.personnel == ParsedSheet(present=True, rows=[])
        assert result.vehicles == ParsedSheet(present=True, rows=[])
        assert result.materials == ParsedSheet(present=True, rows=[])

    def test_error_carries_sheet_and_row(self):
        """The parser knows the cell – the operator has an 18-row sheet and no idea.

        `message` is the line the API hands back, so the sheet name and row number
        have to survive all the way into it.
        """
        file_bytes = create_valid_excel_bytes(
            vehicles=[
                {
                    "name": "TLF 1",
                    "type": "TLF",
                    "display_order": 1,
                    "status": "einsatzbereit",
                    "radio_call_sign": "Florian 1",
                }
            ]
        )
        with pytest.raises(ExcelImportError) as excinfo:
            validate_and_parse_excel(file_bytes)

        error = excinfo.value
        assert error.sheet == "Vehicles"
        assert error.row == 2
        assert error.message.startswith("Vehicles Zeile 2 – ")
        assert "'einsatzbereit'" in error.message

    def test_unreadable_file_does_not_forward_openpyxl_text(self):
        """openpyxl's own wording is the one message here that is not ours to hand out.

        Everything else is a German literal from this module; this site wraps a
        third-party exception, so it gets a fixed sentence and the original goes to
        the log instead.
        """
        with pytest.raises(ExcelImportError) as excinfo:
            validate_and_parse_excel(b"not a valid excel file")

        error = excinfo.value
        assert error.sheet is None and error.row is None
        assert "nicht als Excel-Arbeitsmappe" in error.message
        assert "zip" not in error.message.lower()


# ============================================
# import_data Tests
# ============================================


class TestImportData:
    """Tests for import_data function."""

    @pytest.mark.asyncio
    async def test_import_replace_mode_clears_existing(
        self, db_session: AsyncSession, excel_user: User, sample_personnel: list[Personnel]
    ):
        """Test replace mode deletes all existing data."""
        # Verify existing data
        result = await db_session.execute(select(Personnel))
        assert len(result.scalars().all()) == 3

        # Import new data with replace mode
        parsed_data = full_workbook(
            personnel=[{"name": "New Person", "role": "Test", "status": "available"}], vehicles=[], materials=[]
        )
        counts = await import_data(db_session, parsed_data, "replace", str(excel_user.id))

        assert counts["personnel"] == 1

        # Verify old data is gone and new data is present
        result = await db_session.execute(select(Personnel))
        personnel = result.scalars().all()
        assert len(personnel) == 1
        assert personnel[0].name == "New Person"

    @pytest.mark.asyncio
    async def test_import_append_mode_keeps_existing(
        self, db_session: AsyncSession, excel_user: User, sample_personnel: list[Personnel]
    ):
        """Test append mode keeps existing data."""
        # Verify existing data
        result = await db_session.execute(select(Personnel))
        existing_count = len(result.scalars().all())
        assert existing_count == 3

        # Import new data with append mode
        parsed_data = full_workbook(
            personnel=[{"name": "New Person", "role": "Test", "status": "available"}], vehicles=[], materials=[]
        )
        counts = await import_data(db_session, parsed_data, "append", str(excel_user.id))

        assert counts["personnel"] == 1

        # Verify old data is kept and new data is added
        result = await db_session.execute(select(Personnel))
        personnel = result.scalars().all()
        assert len(personnel) == 4

    @pytest.mark.asyncio
    async def test_import_unknown_mode_does_not_delete(
        self, db_session: AsyncSession, excel_user: User, sample_personnel: list[Personnel]
    ):
        """The removed `merge` mode – and any other stray string – must never delete.

        `merge` used to be a branch here running the same three DELETEs as `replace`. The
        endpoint rejects it now, and this function's fall-through must stay additive so a
        mode that slips past validation can never wipe a station again.
        """
        parsed_data = full_workbook(
            personnel=[{"name": "Merged Person", "role": "Test", "status": "available"}], vehicles=[], materials=[]
        )
        counts = await import_data(db_session, parsed_data, "merge", str(excel_user.id))  # type: ignore[arg-type]

        assert counts["personnel"] == 1
        names = {p.name for p in (await db_session.execute(select(Personnel))).scalars().all()}
        assert "Merged Person" in names
        assert len(names) == len(sample_personnel) + 1

    @pytest.mark.asyncio
    async def test_import_vehicles(self, db_session: AsyncSession, excel_user: User):
        """Test importing vehicles."""
        parsed_data = full_workbook(
            personnel=[],
            vehicles=[
                {
                    "name": "Imported TLF",
                    "type": "TLF",
                    "display_order": 1,
                    "status": "available",
                    "radio_call_sign": "Imported 1",
                }
            ],
            materials=[],
        )
        counts = await import_data(db_session, parsed_data, "replace", str(excel_user.id))

        assert counts["vehicles"] == 1

        result = await db_session.execute(select(Vehicle))
        vehicles = result.scalars().all()
        assert len(vehicles) == 1
        assert vehicles[0].name == "Imported TLF"

    @pytest.mark.asyncio
    async def test_import_materials(self, db_session: AsyncSession, excel_user: User):
        """Test importing materials."""
        parsed_data = full_workbook(
            personnel=[],
            vehicles=[],
            materials=[{"name": "Imported Material", "type": "Pumps", "location": "TLF", "description": "Test"}],
        )
        counts = await import_data(db_session, parsed_data, "replace", str(excel_user.id))

        assert counts["materials"] == 1

        result = await db_session.execute(select(Material))
        materials = result.scalars().all()
        assert len(materials) == 1
        assert materials[0].name == "Imported Material"

    @pytest.mark.asyncio
    async def test_import_returns_counts(self, db_session: AsyncSession, excel_user: User):
        """Test import returns correct counts."""
        parsed_data = full_workbook(
            personnel=[
                {"name": "Person 1", "role": "Test", "status": "available"},
                {"name": "Person 2", "role": "Test", "status": "available"},
            ],
            vehicles=[
                {
                    "name": "Vehicle 1",
                    "type": "TLF",
                    "display_order": 1,
                    "status": "available",
                    "radio_call_sign": "V1",
                }
            ],
            materials=[
                {"name": "Material 1", "type": "Pumps", "location": "TLF"},
                {"name": "Material 2", "type": "Pumps", "location": "TLF"},
                {"name": "Material 3", "type": "Pumps", "location": "TLF"},
            ],
        )
        counts = await import_data(db_session, parsed_data, "replace", str(excel_user.id))

        assert counts["personnel"] == 2
        assert counts["vehicles"] == 1
        assert counts["materials"] == 3

    @pytest.mark.asyncio
    async def test_replace_leaves_tables_whose_sheet_is_absent(
        self,
        db_session: AsyncSession,
        excel_user: User,
        sample_vehicles: list[Vehicle],
        sample_materials: list[Material],
    ):
        """`replace` used to DELETE all three tables whatever the file contained.

        A Personnel-only workbook therefore wiped the fleet and the material list and
        inserted nothing in their place. The endpoint refuses that upload outright; this
        is the second lock, for anyone calling the service directly.
        """
        parsed = ParsedImport(
            personnel=ParsedSheet(present=True, rows=[{"name": "Neu Nina", "role": "Test", "status": "available"}]),
            vehicles=ParsedSheet(present=False),
            materials=ParsedSheet(present=False),
        )
        await import_data(db_session, parsed, "replace", str(excel_user.id))

        assert len((await db_session.execute(select(Vehicle))).scalars().all()) == len(sample_vehicles)
        assert len((await db_session.execute(select(Material))).scalars().all()) == len(sample_materials)

    @pytest.mark.asyncio
    async def test_replace_clears_a_table_whose_sheet_is_present_but_empty(
        self, db_session: AsyncSession, excel_user: User, sample_vehicles: list[Vehicle]
    ):
        """The legitimate case the refusal must not swallow: emptying a table on purpose.

        A sheet with its header and no rows says "the station has no vehicles" – that is
        an answer, and `replace` is allowed to act on it.
        """
        parsed = ParsedImport(
            personnel=ParsedSheet(present=True),
            vehicles=ParsedSheet(present=True),
            materials=ParsedSheet(present=True),
        )
        await import_data(db_session, parsed, "replace", str(excel_user.id))

        assert (await db_session.execute(select(Vehicle))).scalars().all() == []

    @pytest.mark.asyncio
    async def test_import_empty_data(self, db_session: AsyncSession, excel_user: User):
        """Test importing empty data sets."""
        parsed_data = full_workbook(personnel=[], vehicles=[], materials=[])
        counts = await import_data(db_session, parsed_data, "replace", str(excel_user.id))

        assert counts["personnel"] == 0
        assert counts["vehicles"] == 0
        assert counts["materials"] == 0


# ============================================
# export_data_to_excel Tests
# ============================================


class TestExportDataToExcel:
    """Tests for export_data_to_excel function."""

    @pytest.mark.asyncio
    async def test_returns_bytesio(self, db_session: AsyncSession):
        """Test returns a BytesIO buffer."""
        result = await export_data_to_excel(db_session)
        assert isinstance(result, io.BytesIO)

    @pytest.mark.asyncio
    async def test_creates_valid_workbook(self, db_session: AsyncSession):
        """Test creates a valid Excel workbook."""
        result = await export_data_to_excel(db_session)
        wb = load_workbook(result)
        assert wb is not None

    @pytest.mark.asyncio
    async def test_contains_all_sheets(self, db_session: AsyncSession):
        """Test contains all required sheets."""
        result = await export_data_to_excel(db_session)
        wb = load_workbook(result)
        assert "Personnel" in wb.sheetnames
        assert "Vehicles" in wb.sheetnames
        assert "Materials" in wb.sheetnames

    @pytest.mark.asyncio
    async def test_exports_personnel_data(self, db_session: AsyncSession, sample_personnel: list[Personnel]):
        """Test exports personnel data correctly."""
        result = await export_data_to_excel(db_session)
        wb = load_workbook(result)
        ws = wb["Personnel"]

        # Check headers
        headers = [cell.value for cell in ws[1]]
        expected_headers = [col[0] for col in PERSONNEL_COLUMNS]
        assert headers == expected_headers

        # Check data (skip header row)
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Filter out None rows
        data_rows = [row for row in data_rows if row[0] is not None]
        assert len(data_rows) == 3

    @pytest.mark.asyncio
    async def test_exports_vehicles_data(self, db_session: AsyncSession, sample_vehicles: list[Vehicle]):
        """Test exports vehicles data correctly."""
        result = await export_data_to_excel(db_session)
        wb = load_workbook(result)
        ws = wb["Vehicles"]

        # Check headers
        headers = [cell.value for cell in ws[1]]
        expected_headers = [col[0] for col in VEHICLE_COLUMNS]
        assert headers == expected_headers

        # Check data
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        data_rows = [row for row in data_rows if row[0] is not None]
        assert len(data_rows) == 2

    @pytest.mark.asyncio
    async def test_exports_materials_data(self, db_session: AsyncSession, sample_materials: list[Material]):
        """Test exports materials data correctly."""
        result = await export_data_to_excel(db_session)
        wb = load_workbook(result)
        ws = wb["Materials"]

        # Check headers
        headers = [cell.value for cell in ws[1]]
        expected_headers = [col[0] for col in MATERIAL_COLUMNS]
        assert headers == expected_headers

        # Check data
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        data_rows = [row for row in data_rows if row[0] is not None]
        assert len(data_rows) == 3

    @pytest.mark.asyncio
    async def test_exports_empty_database(self, db_session: AsyncSession):
        """Test exports correctly when database is empty."""
        result = await export_data_to_excel(db_session)
        wb = load_workbook(result)

        # Should have sheets but only header rows
        for sheet_name in ["Personnel", "Vehicles", "Materials"]:
            ws = wb[sheet_name]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            data_rows = [row for row in data_rows if row[0] is not None]
            assert len(data_rows) == 0

    @pytest.mark.asyncio
    async def test_buffer_position_at_start(self, db_session: AsyncSession):
        """Test returned buffer is positioned at start."""
        result = await export_data_to_excel(db_session)
        assert result.tell() == 0

    @pytest.mark.asyncio
    async def test_personnel_sorted_by_name(self, db_session: AsyncSession, sample_personnel: list[Personnel]):
        """Test personnel are sorted by name."""
        result = await export_data_to_excel(db_session)
        wb = load_workbook(result)
        ws = wb["Personnel"]

        names = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
        assert names == sorted(names)

    @pytest.mark.asyncio
    async def test_vehicles_sorted_by_display_order(self, db_session: AsyncSession, sample_vehicles: list[Vehicle]):
        """Test vehicles are sorted by display_order."""
        result = await export_data_to_excel(db_session)
        wb = load_workbook(result)
        ws = wb["Vehicles"]

        orders = [row[2] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
        assert orders == sorted(orders)


# ============================================
# Round-Trip Tests
# ============================================


class TestRoundTrip:
    """Tests for export-import round-trip integrity."""

    @pytest.mark.asyncio
    async def test_export_import_round_trip(
        self,
        db_session: AsyncSession,
        excel_user: User,
        sample_personnel: list[Personnel],
        sample_vehicles: list[Vehicle],
        sample_materials: list[Material],
    ):
        """Test data survives export and re-import."""
        # Export current data
        exported = await export_data_to_excel(db_session)

        # Clear database
        from sqlalchemy import delete

        await db_session.execute(delete(Personnel))
        await db_session.execute(delete(Vehicle))
        await db_session.execute(delete(Material))
        await db_session.commit()

        # Re-import
        parsed = validate_and_parse_excel(exported.getvalue())
        counts = await import_data(db_session, parsed, "replace", str(excel_user.id))

        # Verify counts match
        assert counts["personnel"] == len(sample_personnel)
        assert counts["vehicles"] == len(sample_vehicles)
        assert counts["materials"] == len(sample_materials)

    @pytest.mark.asyncio
    async def test_untouched_template_imports_nobody(self, db_session: AsyncSession, excel_user: User):
        """The template goes through the import without adding a single row.

        This test used to assert the opposite – that the examples imported – which is
        precisely how two fictional firefighters, two fictional vehicles and three
        fictional pumps could land on a station's board.
        """
        parsed = validate_and_parse_excel(generate_empty_template().getvalue())

        counts = await import_data(db_session, parsed, "replace", str(excel_user.id))
        assert counts == {"personnel": 0, "vehicles": 0, "materials": 0}

        for model in (Personnel, Vehicle, Material):
            rows = (await db_session.execute(select(model))).scalars().all()
            assert rows == [], f"{model.__name__} got rows from the template"


# ============================================
# Edge Cases
# ============================================


class TestEdgeCases:
    """Tests for edge cases in excel import/export."""

    def test_handles_special_characters(self):
        """Test handles special characters in data."""
        file_bytes = create_valid_excel_bytes(
            personnel=[{"name": "Müller Jürgen", "role": "Führungskraft", "status": "available"}],
            materials=[{"name": "Schläuche", "type": "Schläuche", "location": "TLF", "description": ""}],
        )
        result = validate_and_parse_excel(file_bytes)
        assert result.personnel.rows[0]["name"] == "Müller Jürgen"
        assert result.materials.rows[0]["name"] == "Schläuche"

    def test_handles_very_long_strings(self):
        """Test handles very long text values."""
        long_name = "A" * 1000
        file_bytes = create_valid_excel_bytes(personnel=[{"name": long_name, "role": "Test", "status": "available"}])
        result = validate_and_parse_excel(file_bytes)
        assert result.personnel.rows[0]["name"] == long_name

    @pytest.mark.asyncio
    async def test_import_large_dataset(self, db_session: AsyncSession, excel_user: User):
        """Test importing a large dataset."""
        # Create data for 100 items of each type
        parsed_data = full_workbook(
            personnel=[{"name": f"Person {i}", "role": "Test", "status": "available"} for i in range(100)],
            vehicles=[
                {
                    "name": f"Vehicle {i}",
                    "type": "TLF",
                    "display_order": i,
                    "status": "available",
                    "radio_call_sign": f"V{i}",
                }
                for i in range(100)
            ],
            materials=[{"name": f"Material {i}", "type": "Pumps", "location": "TLF"} for i in range(100)],
        )
        counts = await import_data(db_session, parsed_data, "replace", str(excel_user.id))

        assert counts["personnel"] == 100
        assert counts["vehicles"] == 100
        assert counts["materials"] == 100

    def test_validates_all_personnel_statuses(self):
        """Test all valid personnel statuses are accepted by parser."""
        for status in PERSONNEL_STATUSES:
            file_bytes = create_valid_excel_bytes(personnel=[{"name": "Test", "role": "Test", "status": status}])
            result = validate_and_parse_excel(file_bytes)
            assert result.personnel.rows[0]["status"] == status

    def test_validates_all_vehicle_statuses(self):
        """Test all valid vehicle statuses are accepted."""
        for status in VEHICLE_STATUSES:
            file_bytes = create_valid_excel_bytes(
                vehicles=[
                    {
                        "name": "Test",
                        "type": "TLF",
                        "display_order": 1,
                        "status": status,
                        "radio_call_sign": "Test 1",
                    }
                ]
            )
            result = validate_and_parse_excel(file_bytes)
            assert result.vehicles.rows[0]["status"] == status
