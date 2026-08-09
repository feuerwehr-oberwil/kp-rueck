"""Tests for the Einsätze workbook (plan 25, §7 / decision 21).

The builder is pure and synchronous: it takes an :class:`EventReportData` and
returns the finished XLSX in a buffer, so these tests construct the data in
memory and read the sheet back with openpyxl.

The sheet matches no external format on purpose — it is retyped by hand — so
what is asserted here is what a person reading it while retyping needs: one row
per Schadenplatz including the ones nobody filed a rapport for, a Dauer that
agrees with its own Beginn/Ende, the three-state material answer intact, and
provenance that never claims the KP was the crew.
"""

from datetime import UTC, datetime
from uuid import uuid4

import openpyxl

from app.models import Event, Incident, IncidentAssignment, Personnel, SchadenplatzReport, User
from app.services.audit_export_service import EventReportData
from app.services.excel_import_export import (
    EINSAETZE_COLUMNS,
    build_einsaetze_workbook,
    format_duration,
)

HEADERS = [header for header, _width in EINSAETZE_COLUMNS]


def _sheet(data: EventReportData):
    buffer = build_einsaetze_workbook(data)
    return openpyxl.load_workbook(buffer)["Einsätze"]


def _row(sheet, row_num: int) -> dict[str, object]:
    """One data row as {header: value} so the assertions name their column.

    openpyxl reads an empty cell back as ``None``; the builder writes ``""``.
    Normalised here so "this column is blank" reads the same either way.
    """
    values: dict[str, object] = {}
    for idx, header in enumerate(HEADERS, 1):
        value = sheet.cell(row=row_num, column=idx).value
        values[header] = "" if value is None else value
    return values


def _incident(event: Event, address: str) -> Incident:
    return Incident(
        id=uuid4(),
        event_id=event.id,
        title=address,
        type="elementarereignis",
        priority="medium",
        status="complete",
        location_address=address,
        created_at=datetime(2026, 6, 1, 9, 0, tzinfo=UTC),
    )


def _report(incident_id, **overrides) -> SchadenplatzReport:
    defaults: dict = {
        "id": uuid4(),
        "incident_id": incident_id,
        "is_draft": False,
        "submitted_at": datetime(2026, 6, 1, 12, 32, tzinfo=UTC),
        "created_at": datetime(2026, 6, 1, 12, 0, tzinfo=UTC),
        "updated_at": datetime(2026, 6, 1, 12, 32, tzinfo=UTC),
        "personnel_count_corrected": False,
    }
    defaults.update(overrides)
    return SchadenplatzReport(**defaults)


def _vehicle(name: str, *, present: bool = True) -> dict:
    return {
        "assignment_id": str(uuid4()),
        "vehicle_id": str(uuid4()),
        "name": name,
        "present": present,
    }


def _material(name: str, *, used: bool | None, left_on_site: bool = False, consumable: bool = False) -> dict:
    return {
        "assignment_id": str(uuid4()),
        "material_id": str(uuid4()),
        "name": name,
        "consumable": consumable,
        "used": used,
        "left_on_site": left_on_site,
    }


def _data(event: Event, incidents: list[Incident], reports: list[SchadenplatzReport], **kwargs) -> EventReportData:
    return EventReportData(
        event=event,
        incidents=incidents,
        assignments=kwargs.pop("assignments", []),
        transitions=[],
        reko_reports=[],
        incident_map={inc.id: inc for inc in incidents},
        schadenplatz_reports=reports,
        **kwargs,
    )


class TestSheetShape:
    def test_header_row_is_human_readable(self):
        event = Event(id=uuid4(), name="Sturm 2026", training_flag=False)
        sheet = _sheet(_data(event, [], []))
        assert [c.value for c in sheet[1]] == HEADERS
        assert "Einsatz-Nr." in HEADERS
        assert "Dauer" in HEADERS
        # The sheet does not carry that word any more, in any column.
        assert not any("kostenpflicht" in h.lower() for h in HEADERS)
        assert not any("schadensart" in h.lower() for h in HEADERS)
        # No derived person-hours column (decision 21) — nothing asked for one.
        assert not any("stunden" in h.lower() for h in HEADERS)

    def test_one_row_per_schadenplatz_including_the_ones_without_a_rapport(self):
        """Decision 10: there is no acceptance step, so the gaps must be visible."""
        event = Event(id=uuid4(), name="Sturm 2026", training_flag=False)
        with_rapport = _incident(event, "Bahnhofstrasse 4, Oberwil")
        without_rapport = _incident(event, "Mühlemattstrasse 12, Oberwil")
        report = _report(with_rapport.id, kurzbericht="Keller ausgepumpt.")

        sheet = _sheet(_data(event, [with_rapport, without_rapport], [report]))

        assert sheet.max_row == 3  # header + two Schadenplätze
        first, second = _row(sheet, 2), _row(sheet, 3)
        assert first["Adresse"] == "Bahnhofstrasse 4, Oberwil"
        assert first["Kurzbericht"] == "Keller ausgepumpt."
        # The missing one is a blank row carrying its address, not a missing row.
        assert second["Einsatz-Nr."] == 2
        assert second["Adresse"] == "Mühlemattstrasse 12, Oberwil"
        assert second["Kurzbericht"] == ""
        assert second["Fahrzeuge"] == ""


class TestDuration:
    def test_duration_matches_beginn_and_ende(self):
        event = Event(id=uuid4(), name="Sturm 2026", training_flag=False)
        incident = _incident(event, "Bahnhofstrasse 4, Oberwil")
        report = _report(
            incident.id,
            work_started_at=datetime(2026, 6, 1, 9, 30, tzinfo=UTC),
            work_ended_at=datetime(2026, 6, 1, 11, 10, tzinfo=UTC),
        )
        row = _row(_sheet(_data(event, [incident], [report])), 2)
        assert row["Beginn"] == "01.06.2026 11:30"  # 09:30 UTC = 11:30 CEST
        assert row["Ende"] == "01.06.2026 13:10"
        assert row["Dauer"] == "1:40"

    def test_duration_is_empty_without_both_timestamps(self):
        assert format_duration(None, datetime(2026, 6, 1, 11, 0, tzinfo=UTC)) == ""
        assert format_duration(datetime(2026, 6, 1, 11, 0, tzinfo=UTC), None) == ""

    def test_duration_survives_a_multi_hour_night(self):
        start = datetime(2026, 6, 1, 20, 0, tzinfo=UTC)
        end = datetime(2026, 6, 2, 3, 5, tzinfo=UTC)
        assert format_duration(start, end) == "7:05"


class TestCrewConfirmation:
    def test_corrected_head_count_carries_the_board_value(self):
        """Decision 5: the divergence says the board was behind reality."""
        event = Event(id=uuid4(), name="Sturm 2026", training_flag=False)
        incident = _incident(event, "Bahnhofstrasse 4, Oberwil")
        assignments = [
            IncidentAssignment(
                id=uuid4(),
                incident_id=incident.id,
                resource_type="personnel",
                resource_id=uuid4(),
                assigned_at=datetime(2026, 6, 1, 9, 20, tzinfo=UTC),
            )
            for _ in range(6)
        ]
        report = _report(incident.id, personnel_count=8, personnel_count_corrected=True)

        row = _row(_sheet(_data(event, [incident], [report], assignments=assignments)), 2)
        assert row["Personal"] == 8
        assert row["Personal korrigiert"] == "Ja (Board: 6)"

    def test_fahrzeuge_lists_the_names_the_crew_ticked(self):
        """The list replaced the count: which vehicles, not how many."""
        event = Event(id=uuid4(), name="Sturm 2026", training_flag=False)
        incident = _incident(event, "Bahnhofstrasse 4, Oberwil")
        report = _report(
            incident.id,
            vehicles_json=[
                _vehicle("TLF 1"),
                _vehicle("MTW", present=False),
                _vehicle("Anhänger Pumpe"),
            ],
        )
        row = _row(_sheet(_data(event, [incident], [report])), 2)
        assert row["Fahrzeuge"] == "TLF 1, Anhänger Pumpe"
        # There is no "korrigiert" companion column any more.
        assert "Fahrzeuge korrigiert" not in HEADERS

    def test_fahrzeuge_is_empty_when_the_crew_unticked_everything(self):
        event = Event(id=uuid4(), name="Sturm 2026", training_flag=False)
        incident = _incident(event, "Bahnhofstrasse 4, Oberwil")
        report = _report(incident.id, vehicles_json=[_vehicle("TLF 1", present=False)])
        assert _row(_sheet(_data(event, [incident], [report])), 2)["Fahrzeuge"] == ""


class TestMaterial:
    def test_every_answer_including_keine_angabe_reaches_the_sheet(self):
        event = Event(id=uuid4(), name="Sturm 2026", training_flag=False)
        incident = _incident(event, "Bahnhofstrasse 4, Oberwil")
        report = _report(
            incident.id,
            materials_json=[
                _material("Tauchpumpe", used=True, left_on_site=True),
                _material("Nassauger", used=None),
                _material("Schlauch", used=False),
            ],
            extra_material_note="Pumpe vom Nachbarn",
        )
        row = _row(_sheet(_data(event, [incident], [report])), 2)
        used_cell = str(row["Material gebraucht"])
        assert "Tauchpumpe: gebraucht" in used_cell
        assert "Nassauger: keine Angabe" in used_cell  # decision 14, the third answer
        assert "Schlauch: nicht gebraucht" in used_cell  # decision 16, recorded not acted on
        assert row["Material vor Ort verblieben"] == "Tauchpumpe"
        assert row["Weiteres Material"] == "Pumpe vom Nachbarn"

    def test_consumable_never_gets_a_left_on_site_state(self):
        """Decision 26 — and it must never reach the collect-tomorrow column."""
        event = Event(id=uuid4(), name="Sturm 2026", training_flag=False)
        incident = _incident(event, "Bahnhofstrasse 4, Oberwil")
        report = _report(
            incident.id,
            materials_json=[_material("Ölbindemittel", used=True, left_on_site=True, consumable=True)],
        )
        row = _row(_sheet(_data(event, [incident], [report])), 2)
        assert row["Material gebraucht"] == "Ölbindemittel: gebraucht"
        assert row["Material vor Ort verblieben"] == ""


class TestProvenance:
    def test_field_filed_rapport_says_feld(self):
        event = Event(id=uuid4(), name="Sturm 2026", training_flag=False)
        incident = _incident(event, "Bahnhofstrasse 4, Oberwil")
        person = Personnel(id=uuid4(), name="Muster Hans", role="mannschaft", status="available")
        report = _report(incident.id, created_by_personnel_id=person.id, updated_by_personnel_id=person.id)
        row = _row(
            _sheet(_data(event, [incident], [report], personnel_map={person.id: person})),
            2,
        )
        assert row["Erfasst von"] == "Erfasst von Muster Hans (Feld), 01.06.2026 14:32"

    def test_kp_filed_rapport_says_funkmeldung(self):
        event = Event(id=uuid4(), name="Sturm 2026", training_flag=False)
        incident = _incident(event, "Bahnhofstrasse 4, Oberwil")
        user = User(id=uuid4(), username="beichenberger", display_name="B. Eichenberger", role="editor")
        report = _report(incident.id, created_by_user_id=user.id, updated_by_user_id=user.id)
        row = _row(_sheet(_data(event, [incident], [report], user_map={user.id: user})), 2)
        assert "Erfasst im KP durch B. Eichenberger (Funkmeldung)" in str(row["Erfasst von"])
        assert "(Feld)" not in str(row["Erfasst von"])

    def test_mixed_rapport_shows_both(self):
        event = Event(id=uuid4(), name="Sturm 2026", training_flag=False)
        incident = _incident(event, "Bahnhofstrasse 4, Oberwil")
        person = Personnel(id=uuid4(), name="Muster Hans", role="mannschaft", status="available")
        user = User(id=uuid4(), username="beichenberger", display_name="B. Eichenberger", role="editor")
        report = _report(incident.id, created_by_personnel_id=person.id, updated_by_user_id=user.id)
        row = _row(
            _sheet(
                _data(
                    event,
                    [incident],
                    [report],
                    personnel_map={person.id: person},
                    user_map={user.id: user},
                )
            ),
            2,
        )
        cell = str(row["Erfasst von"])
        assert "Erfasst von Muster Hans (Feld)" in cell
        assert "Zuletzt bearbeitet im KP durch B. Eichenberger (Funkmeldung)" in cell


class TestOwnerBlock:
    def test_owner_and_kfz_land_in_their_own_columns(self):
        event = Event(id=uuid4(), name="Sturm 2026", training_flag=False)
        incident = _incident(event, "Bahnhofstrasse 4, Oberwil")
        report = _report(
            incident.id,
            owner_name="Muster Anna",
            owner_street="Bahnhofstrasse 4",
            owner_city="4104 Oberwil",
            vehicle_plate="BL 123456",
            vehicle_model="VW Golf",
        )
        row = _row(_sheet(_data(event, [incident], [report])), 2)
        assert row["Eigentümer Name"] == "Muster Anna"
        assert row["Eigentümer Strasse"] == "Bahnhofstrasse 4"
        assert row["Eigentümer Ort"] == "4104 Oberwil"
        assert row["KFZ"] == "BL 123456 VW Golf"
