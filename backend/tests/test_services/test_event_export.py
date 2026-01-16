"""Tests for the event_export service module.

Tests ZIP archive generation for event data exports including:
- Event metadata serialization
- Incident data export (JSON + Excel)
- Status transitions export
- Assignments export
- Reko reports export
- Excel summary generation
"""

import io
import json
from datetime import UTC, datetime, timedelta
from uuid import uuid4
from zipfile import ZipFile

import pytest
import pytest_asyncio
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Event,
    Incident,
    IncidentAssignment,
    Personnel,
    RekoReport,
    StatusTransition,
    User,
)
from app.services.event_export import (
    _create_incidents_excel,
    _to_serializable,
    export_event_to_zip,
)


# ============================================
# Fixtures
# ============================================


@pytest_asyncio.fixture
async def export_user(db_session: AsyncSession) -> User:
    """Create a test user for export tests."""
    user = User(
        id=uuid4(),
        username="export_test_user",
        password_hash="$2b$12$test",
        role="editor",
    )
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    return user


@pytest_asyncio.fixture
async def export_event(db_session: AsyncSession) -> Event:
    """Create a test event for export tests."""
    event = Event(
        id=uuid4(),
        name="Export Test Event",
        training_flag=False,
    )
    db_session.add(event)
    await db_session.commit()
    await db_session.refresh(event)
    return event


@pytest_asyncio.fixture
async def export_incident(
    db_session: AsyncSession, export_event: Event, export_user: User
) -> Incident:
    """Create a test incident for export tests."""
    incident = Incident(
        id=uuid4(),
        title="Export Test Incident",
        type="brandbekaempfung",
        priority="high",
        status="einsatz_beendet",
        event_id=export_event.id,
        created_by=export_user.id,
        location_address="Test Location 123",
        location_lat=47.5,
        location_lng=7.5,
        description="Test incident for export",
        completed_at=datetime.now(UTC),
    )
    db_session.add(incident)
    await db_session.commit()
    await db_session.refresh(incident)
    return incident


@pytest_asyncio.fixture
async def export_personnel(db_session: AsyncSession) -> Personnel:
    """Create test personnel for export tests."""
    personnel = Personnel(
        id=uuid4(),
        name="Export Test Person",
        role="Atemschutz",
        availability="available",
    )
    db_session.add(personnel)
    await db_session.commit()
    await db_session.refresh(personnel)
    return personnel


@pytest_asyncio.fixture
async def full_event_with_data(
    db_session: AsyncSession, export_event: Event, export_user: User, export_personnel: Personnel
) -> Event:
    """Create an event with all related data for full export testing."""
    # Create incident
    incident = Incident(
        id=uuid4(),
        title="Full Test Incident",
        type="strassenrettung",
        priority="medium",
        status="disponiert",
        event_id=export_event.id,
        created_by=export_user.id,
        location_address="Full Test Location",
        location_lat=47.6,
        location_lng=7.6,
        description="Full test incident",
    )
    db_session.add(incident)
    await db_session.flush()

    # Create assignment
    assignment = IncidentAssignment(
        incident_id=incident.id,
        resource_type="personnel",
        resource_id=export_personnel.id,
        assigned_by=export_user.id,
    )
    db_session.add(assignment)

    # Create status transition
    transition = StatusTransition(
        incident_id=incident.id,
        from_status="eingegangen",
        to_status="disponiert",
        user_id=export_user.id,
        notes="Status changed for testing",
    )
    db_session.add(transition)

    # Create reko report
    reko = RekoReport(
        incident_id=incident.id,
        token="test-reko-token-123",
        is_relevant=True,
        dangers_json=["fire", "smoke"],
        effort_json={"personnel": 5, "vehicles": 2},
        power_supply="grid",
        summary_text="Test reko summary",
        additional_notes="Additional test notes",
        is_draft=False,
    )
    db_session.add(reko)

    await db_session.commit()
    await db_session.refresh(export_event)
    return export_event


# ============================================
# _to_serializable Tests
# ============================================


class TestToSerializable:
    """Tests for _to_serializable helper function."""

    def test_serializes_datetime(self):
        """Test datetime objects are serialized to ISO format."""
        dt = datetime(2024, 1, 15, 10, 30, 0)
        result = _to_serializable(dt)
        assert result == "2024-01-15T10:30:00"

    def test_serializes_uuid(self):
        """Test UUID objects are serialized to strings."""
        uid = uuid4()
        result = _to_serializable(uid)
        assert result == str(uid)

    def test_serializes_dict(self):
        """Test dictionaries are recursively serialized."""
        uid = uuid4()
        dt = datetime(2024, 1, 15)
        data = {"id": uid, "created": dt, "name": "test"}
        result = _to_serializable(data)
        assert result["id"] == str(uid)
        assert result["created"] == "2024-01-15T00:00:00"
        assert result["name"] == "test"

    def test_serializes_list(self):
        """Test lists are recursively serialized."""
        uid1 = uuid4()
        uid2 = uuid4()
        data = [uid1, uid2, "string"]
        result = _to_serializable(data)
        assert result == [str(uid1), str(uid2), "string"]

    def test_serializes_tuple(self):
        """Test tuples are serialized as lists."""
        uid = uuid4()
        data = (uid, "test")
        result = _to_serializable(data)
        assert result == [str(uid), "test"]

    def test_returns_primitives_unchanged(self):
        """Test primitive types are returned unchanged."""
        assert _to_serializable("string") == "string"
        assert _to_serializable(123) == 123
        assert _to_serializable(45.67) == 45.67
        assert _to_serializable(True) is True
        assert _to_serializable(None) is None

    def test_serializes_nested_structure(self):
        """Test deeply nested structures are fully serialized."""
        uid = uuid4()
        dt = datetime(2024, 1, 15)
        data = {"outer": {"inner": [{"id": uid, "date": dt}]}}
        result = _to_serializable(data)
        assert result["outer"]["inner"][0]["id"] == str(uid)
        assert result["outer"]["inner"][0]["date"] == "2024-01-15T00:00:00"


# ============================================
# export_event_to_zip Tests
# ============================================


class TestExportEventToZip:
    """Tests for export_event_to_zip main function."""

    @pytest.mark.asyncio
    async def test_raises_error_for_nonexistent_event(self, db_session: AsyncSession):
        """Test raises ValueError for non-existent event."""
        fake_id = str(uuid4())
        with pytest.raises(ValueError, match=f"Event {fake_id} not found"):
            await export_event_to_zip(db_session, fake_id)

    @pytest.mark.asyncio
    async def test_returns_bytesio_buffer(self, db_session: AsyncSession, export_event: Event):
        """Test returns a BytesIO buffer."""
        result = await export_event_to_zip(db_session, str(export_event.id))
        assert isinstance(result, io.BytesIO)

    @pytest.mark.asyncio
    async def test_creates_valid_zip_archive(self, db_session: AsyncSession, export_event: Event):
        """Test creates a valid ZIP archive."""
        result = await export_event_to_zip(db_session, str(export_event.id))
        with ZipFile(result, "r") as zf:
            # Should not raise
            zf.testzip()

    @pytest.mark.asyncio
    async def test_zip_contains_required_files(
        self, db_session: AsyncSession, export_event: Event
    ):
        """Test ZIP archive contains all required files."""
        result = await export_event_to_zip(db_session, str(export_event.id))
        with ZipFile(result, "r") as zf:
            file_names = zf.namelist()
            assert "event_metadata.json" in file_names
            assert "incidents.json" in file_names
            assert "assignments.json" in file_names
            assert "status_transitions.json" in file_names
            assert "reko_reports.json" in file_names
            assert "incidents_summary.xlsx" in file_names
            assert "README.txt" in file_names

    @pytest.mark.asyncio
    async def test_event_metadata_json_valid(
        self, db_session: AsyncSession, export_event: Event
    ):
        """Test event metadata JSON is valid and contains correct data."""
        result = await export_event_to_zip(db_session, str(export_event.id))
        with ZipFile(result, "r") as zf:
            metadata_content = zf.read("event_metadata.json").decode("utf-8")
            metadata = json.loads(metadata_content)

            assert metadata["id"] == str(export_event.id)
            assert metadata["name"] == export_event.name
            assert metadata["training_flag"] == export_event.training_flag
            assert "created_at" in metadata
            assert "updated_at" in metadata
            assert "incident_count" in metadata

    @pytest.mark.asyncio
    async def test_incidents_json_includes_incident_data(
        self, db_session: AsyncSession, export_event: Event, export_incident: Incident
    ):
        """Test incidents JSON includes all incident data."""
        result = await export_event_to_zip(db_session, str(export_event.id))
        with ZipFile(result, "r") as zf:
            incidents_content = zf.read("incidents.json").decode("utf-8")
            incidents = json.loads(incidents_content)

            assert len(incidents) == 1
            inc = incidents[0]
            assert inc["id"] == str(export_incident.id)
            assert inc["title"] == export_incident.title
            assert inc["type"] == export_incident.type
            assert inc["priority"] == export_incident.priority
            assert inc["status"] == export_incident.status
            assert inc["location_address"] == export_incident.location_address
            assert inc["description"] == export_incident.description

    @pytest.mark.asyncio
    async def test_full_event_export_includes_all_related_data(
        self, db_session: AsyncSession, full_event_with_data: Event
    ):
        """Test full export includes incidents, assignments, transitions, and reko reports."""
        result = await export_event_to_zip(db_session, str(full_event_with_data.id))
        with ZipFile(result, "r") as zf:
            # Check incidents
            incidents = json.loads(zf.read("incidents.json").decode("utf-8"))
            assert len(incidents) >= 1

            # Check assignments
            assignments = json.loads(zf.read("assignments.json").decode("utf-8"))
            assert len(assignments) >= 1
            assert "resource_type" in assignments[0]
            assert "resource_id" in assignments[0]

            # Check status transitions
            transitions = json.loads(zf.read("status_transitions.json").decode("utf-8"))
            assert len(transitions) >= 1
            assert "from_status" in transitions[0]
            assert "to_status" in transitions[0]

            # Check reko reports
            reko_reports = json.loads(zf.read("reko_reports.json").decode("utf-8"))
            assert len(reko_reports) >= 1
            assert "is_relevant" in reko_reports[0]
            assert "summary_text" in reko_reports[0]

    @pytest.mark.asyncio
    async def test_readme_contains_event_info(
        self, db_session: AsyncSession, export_event: Event
    ):
        """Test README contains event information."""
        result = await export_event_to_zip(db_session, str(export_event.id))
        with ZipFile(result, "r") as zf:
            readme_content = zf.read("README.txt").decode("utf-8")
            assert export_event.name in readme_content
            assert str(export_event.id) in readme_content
            assert "Export Date" in readme_content
            assert "Files Included" in readme_content

    @pytest.mark.asyncio
    async def test_excel_file_is_valid(
        self, db_session: AsyncSession, export_event: Event, export_incident: Incident
    ):
        """Test Excel summary file is valid and contains data."""
        result = await export_event_to_zip(db_session, str(export_event.id))
        with ZipFile(result, "r") as zf:
            excel_content = zf.read("incidents_summary.xlsx")
            excel_buffer = io.BytesIO(excel_content)
            wb = load_workbook(excel_buffer)

            assert "Incidents" in wb.sheetnames
            ws = wb["Incidents"]

            # Check header row
            assert ws.cell(row=1, column=1).value == "ID"
            assert ws.cell(row=1, column=2).value == "Title"

            # Check data row
            assert ws.cell(row=2, column=2).value == export_incident.title

    @pytest.mark.asyncio
    async def test_handles_incident_without_optional_fields(
        self, db_session: AsyncSession, export_event: Event, export_user: User
    ):
        """Test export handles incidents with null optional fields."""
        # Create minimal incident without optional fields
        incident = Incident(
            id=uuid4(),
            title="Minimal Incident",
            type="brandbekaempfung",
            priority="low",
            status="eingegangen",
            event_id=export_event.id,
            created_by=export_user.id,
            # No location, description, completed_at
        )
        db_session.add(incident)
        await db_session.commit()

        result = await export_event_to_zip(db_session, str(export_event.id))
        with ZipFile(result, "r") as zf:
            incidents = json.loads(zf.read("incidents.json").decode("utf-8"))
            minimal_inc = next(i for i in incidents if i["title"] == "Minimal Incident")
            assert minimal_inc["location_lat"] is None
            assert minimal_inc["location_lng"] is None
            assert minimal_inc["completed_at"] is None

    @pytest.mark.asyncio
    async def test_handles_training_event(
        self, db_session: AsyncSession
    ):
        """Test export works for training mode events."""
        training_event = Event(
            id=uuid4(),
            name="Training Event",
            training_flag=True,
        )
        db_session.add(training_event)
        await db_session.commit()

        result = await export_event_to_zip(db_session, str(training_event.id))
        with ZipFile(result, "r") as zf:
            metadata = json.loads(zf.read("event_metadata.json").decode("utf-8"))
            assert metadata["training_flag"] is True

    @pytest.mark.asyncio
    async def test_handles_archived_event(
        self, db_session: AsyncSession
    ):
        """Test export includes archived_at timestamp when present."""
        archived_event = Event(
            id=uuid4(),
            name="Archived Event",
            training_flag=False,
            archived_at=datetime.now(UTC) - timedelta(days=1),
        )
        db_session.add(archived_event)
        await db_session.commit()

        result = await export_event_to_zip(db_session, str(archived_event.id))
        with ZipFile(result, "r") as zf:
            metadata = json.loads(zf.read("event_metadata.json").decode("utf-8"))
            assert metadata["archived_at"] is not None

    @pytest.mark.asyncio
    async def test_json_uses_utf8_encoding(
        self, db_session: AsyncSession, export_event: Event, export_user: User
    ):
        """Test JSON files use UTF-8 encoding for special characters."""
        # Create incident with German special characters
        incident = Incident(
            id=uuid4(),
            title="Übung mit Säure",
            type="brandbekaempfung",
            priority="high",
            status="eingegangen",
            event_id=export_event.id,
            created_by=export_user.id,
            location_address="Mühlstraße 42, Zürich",
            description="Gefährliche Stoffe: Ätzend",
        )
        db_session.add(incident)
        await db_session.commit()

        result = await export_event_to_zip(db_session, str(export_event.id))
        with ZipFile(result, "r") as zf:
            incidents = json.loads(zf.read("incidents.json").decode("utf-8"))
            special_inc = next(i for i in incidents if "Übung" in i["title"])
            assert "Übung mit Säure" in special_inc["title"]
            assert "Mühlstraße" in special_inc["location_address"]
            assert "Ätzend" in special_inc["description"]


# ============================================
# _create_incidents_excel Tests
# ============================================


class TestCreateIncidentsExcel:
    """Tests for _create_incidents_excel helper function."""

    def test_creates_valid_workbook(self, export_event: Event):
        """Test creates a valid Excel workbook."""
        incidents_data = []
        result = _create_incidents_excel(export_event, incidents_data)

        wb = load_workbook(result)
        assert "Incidents" in wb.sheetnames

    def test_includes_headers(self, export_event: Event):
        """Test Excel has correct headers."""
        incidents_data = []
        result = _create_incidents_excel(export_event, incidents_data)

        wb = load_workbook(result)
        ws = wb["Incidents"]

        expected_headers = [
            "ID",
            "Title",
            "Type",
            "Priority",
            "Status",
            "Location",
            "Description",
            "Created At",
            "Completed At",
        ]
        for col, expected in enumerate(expected_headers, 1):
            assert ws.cell(row=1, column=col).value == expected

    def test_includes_incident_data(self, export_event: Event):
        """Test Excel includes incident data rows."""
        incidents_data = [
            {
                "id": str(uuid4()),
                "title": "Test Incident",
                "type": "brandbekaempfung",
                "priority": "high",
                "status": "eingegangen",
                "location_address": "Test Location",
                "description": "Test description",
                "created_at": "2024-01-15T10:00:00",
                "completed_at": None,
            }
        ]
        result = _create_incidents_excel(export_event, incidents_data)

        wb = load_workbook(result)
        ws = wb["Incidents"]

        assert ws.cell(row=2, column=2).value == "Test Incident"
        assert ws.cell(row=2, column=3).value == "brandbekaempfung"
        assert ws.cell(row=2, column=4).value == "high"
        assert ws.cell(row=2, column=6).value == "Test Location"

    def test_handles_multiple_incidents(self, export_event: Event):
        """Test Excel handles multiple incidents."""
        incidents_data = [
            {
                "id": str(uuid4()),
                "title": f"Incident {i}",
                "type": "brandbekaempfung",
                "priority": "medium",
                "status": "eingegangen",
                "location_address": f"Location {i}",
                "description": f"Description {i}",
                "created_at": "2024-01-15T10:00:00",
                "completed_at": None,
            }
            for i in range(5)
        ]
        result = _create_incidents_excel(export_event, incidents_data)

        wb = load_workbook(result)
        ws = wb["Incidents"]

        # Check all 5 incidents are present (plus header row)
        for i in range(5):
            assert ws.cell(row=i + 2, column=2).value == f"Incident {i}"

    def test_handles_empty_incidents(self, export_event: Event):
        """Test Excel handles empty incidents list."""
        incidents_data = []
        result = _create_incidents_excel(export_event, incidents_data)

        wb = load_workbook(result)
        ws = wb["Incidents"]

        # Should have headers but no data rows
        assert ws.cell(row=1, column=1).value == "ID"
        assert ws.cell(row=2, column=1).value is None

    def test_handles_none_values(self, export_event: Event):
        """Test Excel handles None values gracefully."""
        incidents_data = [
            {
                "id": str(uuid4()),
                "title": "Test",
                "type": "brandbekaempfung",
                "priority": "low",
                "status": "eingegangen",
                "location_address": None,
                "description": None,
                "created_at": "2024-01-15T10:00:00",
                "completed_at": None,
            }
        ]
        result = _create_incidents_excel(export_event, incidents_data)

        wb = load_workbook(result)
        ws = wb["Incidents"]

        assert ws.cell(row=2, column=6).value is None  # location_address
        assert ws.cell(row=2, column=7).value is None  # description

    def test_header_styling_applied(self, export_event: Event):
        """Test header cells have styling applied."""
        incidents_data = []
        result = _create_incidents_excel(export_event, incidents_data)

        wb = load_workbook(result)
        ws = wb["Incidents"]

        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True

    def test_returns_bytesio_at_start(self, export_event: Event):
        """Test returned buffer is at position 0."""
        incidents_data = []
        result = _create_incidents_excel(export_event, incidents_data)

        assert result.tell() == 0  # Position should be at start


# ============================================
# Edge Cases
# ============================================


class TestExportEdgeCases:
    """Tests for edge cases in event export."""

    @pytest.mark.asyncio
    async def test_event_with_many_incidents(
        self, db_session: AsyncSession, export_event: Event, export_user: User
    ):
        """Test export handles event with many incidents."""
        # Create 20 incidents
        for i in range(20):
            incident = Incident(
                id=uuid4(),
                title=f"Bulk Incident {i}",
                type="brandbekaempfung",
                priority="low",
                status="eingegangen",
                event_id=export_event.id,
                created_by=export_user.id,
            )
            db_session.add(incident)
        await db_session.commit()

        result = await export_event_to_zip(db_session, str(export_event.id))
        with ZipFile(result, "r") as zf:
            incidents = json.loads(zf.read("incidents.json").decode("utf-8"))
            assert len(incidents) == 20

            # Check Excel too
            excel_content = zf.read("incidents_summary.xlsx")
            excel_buffer = io.BytesIO(excel_content)
            wb = load_workbook(excel_buffer)
            ws = wb["Incidents"]
            # Count non-empty rows (excluding header)
            data_rows = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value)
            assert data_rows == 20

    @pytest.mark.asyncio
    async def test_incident_with_long_description(
        self, db_session: AsyncSession, export_event: Event, export_user: User
    ):
        """Test export handles very long descriptions."""
        long_description = "A" * 10000  # 10,000 characters
        incident = Incident(
            id=uuid4(),
            title="Long Description Incident",
            type="brandbekaempfung",
            priority="medium",
            status="eingegangen",
            event_id=export_event.id,
            created_by=export_user.id,
            description=long_description,
        )
        db_session.add(incident)
        await db_session.commit()

        result = await export_event_to_zip(db_session, str(export_event.id))
        with ZipFile(result, "r") as zf:
            incidents = json.loads(zf.read("incidents.json").decode("utf-8"))
            long_inc = next(i for i in incidents if i["title"] == "Long Description Incident")
            assert len(long_inc["description"]) == 10000

    @pytest.mark.asyncio
    async def test_reko_with_complex_json_fields(
        self, db_session: AsyncSession, export_event: Event, export_user: User
    ):
        """Test export handles reko reports with complex JSON fields."""
        incident = Incident(
            id=uuid4(),
            title="Complex Reko Incident",
            type="brandbekaempfung",
            priority="high",
            status="eingegangen",
            event_id=export_event.id,
            created_by=export_user.id,
        )
        db_session.add(incident)
        await db_session.flush()

        # Create reko with complex nested JSON
        reko = RekoReport(
            incident_id=incident.id,
            token="complex-reko-token-456",
            is_relevant=True,
            dangers_json=["fire", "smoke", "chemical", "structural_collapse"],
            effort_json={
                "personnel": {"firefighters": 10, "paramedics": 4},
                "vehicles": {"tlf": 2, "dlk": 1, "rtw": 2},
                "estimated_duration_hours": 6,
            },
            photos_json=[
                {"id": "photo1", "url": "/photos/1.jpg"},
                {"id": "photo2", "url": "/photos/2.jpg"},
            ],
            power_supply="generator",
            summary_text="Complex test summary",
            is_draft=False,
        )
        db_session.add(reko)
        await db_session.commit()

        result = await export_event_to_zip(db_session, str(export_event.id))
        with ZipFile(result, "r") as zf:
            reko_reports = json.loads(zf.read("reko_reports.json").decode("utf-8"))
            assert len(reko_reports) >= 1
            report = reko_reports[0]
            assert isinstance(report["dangers_json"], list)
            assert isinstance(report["effort_json"], dict)
            assert isinstance(report["photos_json"], list)
