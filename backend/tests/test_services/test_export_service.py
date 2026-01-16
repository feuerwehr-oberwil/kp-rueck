"""Tests for export_service.py - Event export functionality."""

from datetime import UTC
from io import BytesIO
from uuid import uuid4

import pytest
import pytest_asyncio
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Event,
    Incident,
    IncidentAssignment,
    Material,
    Personnel,
    RekoReport,
    StatusTransition,
    User,
    Vehicle,
)
from app.services.export_service import (
    export_event_excel,
    export_event_pdf,
    export_event_photos,
)

# ============================================
# Fixtures
# ============================================


@pytest_asyncio.fixture
async def export_user(db_session: AsyncSession) -> User:
    """Create a user for export tests."""
    user = User(
        id=uuid4(),
        username="export_test_user",
        password_hash="hashed_password",
        role="editor",
    )
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    return user


@pytest_asyncio.fixture
async def export_event(db_session: AsyncSession) -> Event:
    """Create an event for export tests."""
    event = Event(
        id=uuid4(),
        name="Test Export Event",
        training_flag=False,
    )
    db_session.add(event)
    await db_session.commit()
    await db_session.refresh(event)
    return event


@pytest_asyncio.fixture
async def export_training_event(db_session: AsyncSession) -> Event:
    """Create a training event for export tests."""
    event = Event(
        id=uuid4(),
        name="Training Export Event",
        training_flag=True,
    )
    db_session.add(event)
    await db_session.commit()
    await db_session.refresh(event)
    return event


@pytest_asyncio.fixture
async def export_incident(db_session: AsyncSession, export_event: Event, export_user: User) -> Incident:
    """Create an incident for export tests."""
    incident = Incident(
        id=uuid4(),
        title="Test Export Incident",
        type="brandbekaempfung",
        priority="high",
        status="eingegangen",
        location_address="Test Street 123, Basel",
        location_lat=47.5596,
        location_lng=7.5886,
        description="Test incident description for export",
        event_id=export_event.id,
        created_by=export_user.id,
    )
    db_session.add(incident)
    await db_session.commit()
    await db_session.refresh(incident)
    return incident


@pytest_asyncio.fixture
async def export_incident_completed(db_session: AsyncSession, export_event: Event, export_user: User) -> Incident:
    """Create a completed incident for export tests."""
    from datetime import datetime, timedelta

    incident = Incident(
        id=uuid4(),
        title="Completed Export Incident",
        type="technische_hilfeleistung",
        priority="medium",
        status="abschluss",  # Valid status: eingegangen, reko, disponiert, einsatz, einsatz_beendet, abschluss
        location_address="Another Street 456, Basel",
        location_lat=47.5600,
        location_lng=7.5900,
        description="Completed incident for export",
        event_id=export_event.id,
        created_by=export_user.id,
        completed_at=datetime.now(UTC) + timedelta(hours=2),
    )
    db_session.add(incident)
    await db_session.commit()
    await db_session.refresh(incident)
    return incident


@pytest_asyncio.fixture
async def export_personnel(db_session: AsyncSession) -> Personnel:
    """Create personnel for export tests."""
    personnel = Personnel(
        id=uuid4(),
        name="Export Test Person",
        role="Gruppenführer",
        availability="available",
    )
    db_session.add(personnel)
    await db_session.commit()
    await db_session.refresh(personnel)
    return personnel


@pytest_asyncio.fixture
async def export_vehicle(db_session: AsyncSession) -> Vehicle:
    """Create vehicle for export tests."""
    vehicle = Vehicle(
        id=uuid4(),
        name="TLF Export",
        type="TLF",
        status="available",
    )
    db_session.add(vehicle)
    await db_session.commit()
    await db_session.refresh(vehicle)
    return vehicle


@pytest_asyncio.fixture
async def export_material(db_session: AsyncSession) -> Material:
    """Create material for export tests."""
    material = Material(
        id=uuid4(),
        name="Export Stromerzeuger",
        type="Stromerzeuger",
        status="available",
        location="Lager",
    )
    db_session.add(material)
    await db_session.commit()
    await db_session.refresh(material)
    return material


@pytest_asyncio.fixture
async def export_assignment_personnel(
    db_session: AsyncSession, export_incident: Incident, export_personnel: Personnel
) -> IncidentAssignment:
    """Create a personnel assignment for export tests."""
    assignment = IncidentAssignment(
        id=uuid4(),
        incident_id=export_incident.id,
        resource_type="personnel",
        resource_id=export_personnel.id,
    )
    db_session.add(assignment)
    await db_session.commit()
    await db_session.refresh(assignment)
    return assignment


@pytest_asyncio.fixture
async def export_assignment_vehicle(
    db_session: AsyncSession, export_incident: Incident, export_vehicle: Vehicle
) -> IncidentAssignment:
    """Create a vehicle assignment for export tests."""
    assignment = IncidentAssignment(
        id=uuid4(),
        incident_id=export_incident.id,
        resource_type="vehicle",
        resource_id=export_vehicle.id,
    )
    db_session.add(assignment)
    await db_session.commit()
    await db_session.refresh(assignment)
    return assignment


@pytest_asyncio.fixture
async def export_assignment_material(
    db_session: AsyncSession, export_incident: Incident, export_material: Material
) -> IncidentAssignment:
    """Create a material assignment for export tests."""
    assignment = IncidentAssignment(
        id=uuid4(),
        incident_id=export_incident.id,
        resource_type="material",
        resource_id=export_material.id,
    )
    db_session.add(assignment)
    await db_session.commit()
    await db_session.refresh(assignment)
    return assignment


@pytest_asyncio.fixture
async def export_status_transition(
    db_session: AsyncSession, export_incident: Incident, export_user: User
) -> StatusTransition:
    """Create a status transition for export tests."""
    transition = StatusTransition(
        id=uuid4(),
        incident_id=export_incident.id,
        from_status="eingegangen",
        to_status="reko",  # Valid status
        user_id=export_user.id,  # Changed from changed_by
    )
    db_session.add(transition)
    await db_session.commit()
    await db_session.refresh(transition)
    return transition


@pytest_asyncio.fixture
async def export_reko_report(db_session: AsyncSession, export_incident: Incident) -> RekoReport:
    """Create a Reko report for export tests."""
    import secrets

    report = RekoReport(
        id=uuid4(),
        incident_id=export_incident.id,
        token=secrets.token_urlsafe(32),  # Required field
        is_relevant=True,
        summary_text="Test Reko summary for export",
        additional_notes="Additional test notes",
        is_draft=False,
    )
    db_session.add(report)
    await db_session.commit()
    await db_session.refresh(report)
    return report


# ============================================
# PDF Export Tests
# ============================================


class TestExportEventPdf:
    """Tests for export_event_pdf function."""

    @pytest.mark.asyncio
    async def test_export_pdf_basic_event(self, db_session: AsyncSession, export_event: Event, export_user: User):
        """Test PDF export with basic event (no incidents)."""
        result = await export_event_pdf(db_session, export_event.id, export_user)

        assert result is not None
        assert isinstance(result, BytesIO)
        # PDF should have content
        content = result.read()
        assert len(content) > 0
        # PDF magic bytes
        assert content[:4] == b"%PDF"

    @pytest.mark.asyncio
    async def test_export_pdf_with_incident(
        self,
        db_session: AsyncSession,
        export_event: Event,
        export_incident: Incident,
        export_user: User,
    ):
        """Test PDF export with an incident."""
        result = await export_event_pdf(db_session, export_event.id, export_user)

        assert result is not None
        content = result.read()
        assert len(content) > 0
        assert content[:4] == b"%PDF"

    @pytest.mark.asyncio
    async def test_export_pdf_with_completed_incident(
        self,
        db_session: AsyncSession,
        export_event: Event,
        export_incident_completed: Incident,
        export_user: User,
    ):
        """Test PDF export with a completed incident."""
        result = await export_event_pdf(db_session, export_event.id, export_user)

        assert result is not None
        content = result.read()
        assert content[:4] == b"%PDF"

    @pytest.mark.asyncio
    async def test_export_pdf_with_assignments(
        self,
        db_session: AsyncSession,
        export_event: Event,
        export_incident: Incident,
        export_assignment_personnel: IncidentAssignment,
        export_assignment_vehicle: IncidentAssignment,
        export_assignment_material: IncidentAssignment,
        export_user: User,
    ):
        """Test PDF export with personnel, vehicle, and material assignments."""
        result = await export_event_pdf(db_session, export_event.id, export_user)

        assert result is not None
        content = result.read()
        assert content[:4] == b"%PDF"

    @pytest.mark.asyncio
    async def test_export_pdf_with_status_transitions(
        self,
        db_session: AsyncSession,
        export_event: Event,
        export_incident: Incident,
        export_status_transition: StatusTransition,
        export_user: User,
    ):
        """Test PDF export with status transitions."""
        result = await export_event_pdf(db_session, export_event.id, export_user)

        assert result is not None
        content = result.read()
        assert content[:4] == b"%PDF"

    @pytest.mark.asyncio
    async def test_export_pdf_with_reko_report(
        self,
        db_session: AsyncSession,
        export_event: Event,
        export_incident: Incident,
        export_reko_report: RekoReport,
        export_user: User,
    ):
        """Test PDF export with Reko report."""
        result = await export_event_pdf(db_session, export_event.id, export_user)

        assert result is not None
        content = result.read()
        assert content[:4] == b"%PDF"

    @pytest.mark.asyncio
    async def test_export_pdf_training_event(
        self, db_session: AsyncSession, export_training_event: Event, export_user: User
    ):
        """Test PDF export with training flag enabled."""
        result = await export_event_pdf(db_session, export_training_event.id, export_user)

        assert result is not None
        content = result.read()
        assert content[:4] == b"%PDF"

    @pytest.mark.asyncio
    async def test_export_pdf_invalid_event_id(self, db_session: AsyncSession, export_user: User):
        """Test PDF export with non-existent event ID."""
        fake_id = uuid4()

        with pytest.raises(ValueError) as exc_info:
            await export_event_pdf(db_session, fake_id, export_user)

        assert str(fake_id) in str(exc_info.value)
        assert "not found" in str(exc_info.value)

    @pytest.mark.asyncio
    async def test_export_pdf_excludes_deleted_incidents(
        self,
        db_session: AsyncSession,
        export_event: Event,
        export_incident: Incident,
        export_user: User,
    ):
        """Test that PDF export excludes soft-deleted incidents."""
        from datetime import datetime

        # Soft delete the incident
        export_incident.deleted_at = datetime.now(UTC)
        await db_session.commit()

        result = await export_event_pdf(db_session, export_event.id, export_user)

        # Should still generate PDF but without the deleted incident
        assert result is not None
        content = result.read()
        assert content[:4] == b"%PDF"


# ============================================
# Excel Export Tests
# ============================================


class TestExportEventExcel:
    """Tests for export_event_excel function."""

    @pytest.mark.asyncio
    async def test_export_excel_basic_event(self, db_session: AsyncSession, export_event: Event):
        """Test Excel export with basic event (no incidents)."""
        result = await export_event_excel(db_session, export_event.id)

        assert result is not None
        assert isinstance(result, BytesIO)

        # Load and verify workbook
        wb = load_workbook(result)
        assert "Übersicht" in wb.sheetnames
        assert "Einsätze" in wb.sheetnames
        assert "Einsatz Details" in wb.sheetnames

        # Verify overview sheet
        ws_overview = wb["Übersicht"]
        assert ws_overview["A1"].value == "Einsatzbericht"
        assert ws_overview["B3"].value == export_event.name

    @pytest.mark.asyncio
    async def test_export_excel_with_incident(
        self, db_session: AsyncSession, export_event: Event, export_incident: Incident
    ):
        """Test Excel export with an incident."""
        result = await export_event_excel(db_session, export_event.id)

        wb = load_workbook(result)

        # Verify incidents sheet has data
        ws_incidents = wb["Einsätze"]
        # Header row
        assert ws_incidents.cell(row=1, column=1).value == "Nr"
        assert ws_incidents.cell(row=1, column=2).value == "Titel"
        # Data row
        assert ws_incidents.cell(row=2, column=1).value == 1
        assert ws_incidents.cell(row=2, column=2).value == export_incident.title

    @pytest.mark.asyncio
    async def test_export_excel_with_completed_incident(
        self, db_session: AsyncSession, export_event: Event, export_incident_completed: Incident
    ):
        """Test Excel export with completed incident shows completion date."""
        result = await export_event_excel(db_session, export_event.id)

        wb = load_workbook(result)
        ws_incidents = wb["Einsätze"]

        # Completion date should be present (column 8)
        completion_cell = ws_incidents.cell(row=2, column=8).value
        assert completion_cell is not None
        assert completion_cell != ""

    @pytest.mark.asyncio
    async def test_export_excel_with_assignments(
        self,
        db_session: AsyncSession,
        export_event: Event,
        export_incident: Incident,
        export_assignment_personnel: IncidentAssignment,
        export_assignment_vehicle: IncidentAssignment,
        export_personnel: Personnel,
        export_vehicle: Vehicle,
    ):
        """Test Excel export includes assignments in details sheet."""
        result = await export_event_excel(db_session, export_event.id)

        wb = load_workbook(result)
        ws_details = wb["Einsatz Details"]

        # Details sheet should have assignment info
        # Just verify the sheet was created and has content
        assert ws_details.cell(row=1, column=1).value is not None

    @pytest.mark.asyncio
    async def test_export_excel_with_status_transitions(
        self,
        db_session: AsyncSession,
        export_event: Event,
        export_incident: Incident,
        export_status_transition: StatusTransition,
    ):
        """Test Excel export includes status transitions."""
        result = await export_event_excel(db_session, export_event.id)

        wb = load_workbook(result)
        ws_details = wb["Einsatz Details"]

        # Verify sheet has content
        assert ws_details.cell(row=1, column=1).value is not None

    @pytest.mark.asyncio
    async def test_export_excel_invalid_event_id(self, db_session: AsyncSession):
        """Test Excel export with non-existent event ID."""
        fake_id = uuid4()

        with pytest.raises(ValueError) as exc_info:
            await export_event_excel(db_session, fake_id)

        assert str(fake_id) in str(exc_info.value)
        assert "not found" in str(exc_info.value)

    @pytest.mark.asyncio
    async def test_export_excel_training_mode_display(self, db_session: AsyncSession, export_training_event: Event):
        """Test Excel export correctly displays training mode."""
        result = await export_event_excel(db_session, export_training_event.id)

        wb = load_workbook(result)
        ws_overview = wb["Übersicht"]

        # Training mode cell should show "Ja"
        assert ws_overview["B6"].value == "Ja"

    @pytest.mark.asyncio
    async def test_export_excel_non_training_mode_display(self, db_session: AsyncSession, export_event: Event):
        """Test Excel export correctly displays non-training mode."""
        result = await export_event_excel(db_session, export_event.id)

        wb = load_workbook(result)
        ws_overview = wb["Übersicht"]

        # Training mode cell should show "Nein"
        assert ws_overview["B6"].value == "Nein"

    @pytest.mark.asyncio
    async def test_export_excel_excludes_deleted_incidents(
        self,
        db_session: AsyncSession,
        export_event: Event,
        export_incident: Incident,
    ):
        """Test that Excel export excludes soft-deleted incidents."""
        from datetime import datetime

        # Soft delete the incident
        export_incident.deleted_at = datetime.now(UTC)
        await db_session.commit()

        result = await export_event_excel(db_session, export_event.id)

        wb = load_workbook(result)
        ws_incidents = wb["Einsätze"]

        # Should only have header row, no data rows
        assert ws_incidents.cell(row=2, column=1).value is None

    @pytest.mark.asyncio
    async def test_export_excel_multiple_incidents(
        self, db_session: AsyncSession, export_event: Event, export_user: User
    ):
        """Test Excel export with multiple incidents."""
        # Create additional incidents
        for i in range(3):
            incident = Incident(
                id=uuid4(),
                title=f"Multi Export Incident {i + 1}",
                type="brandbekaempfung",
                priority="medium",
                status="eingegangen",
                event_id=export_event.id,
                created_by=export_user.id,
            )
            db_session.add(incident)
        await db_session.commit()

        result = await export_event_excel(db_session, export_event.id)

        wb = load_workbook(result)
        ws_incidents = wb["Einsätze"]

        # Should have 3 data rows
        assert ws_incidents.cell(row=2, column=1).value == 1
        assert ws_incidents.cell(row=3, column=1).value == 2
        assert ws_incidents.cell(row=4, column=1).value == 3


# ============================================
# Photos Export Tests
# ============================================


class TestExportEventPhotos:
    """Tests for export_event_photos function."""

    @pytest.mark.asyncio
    async def test_export_photos_no_photos(self, db_session: AsyncSession, export_event: Event):
        """Test photos export with no photos returns None."""
        result = await export_event_photos(db_session, export_event.id)

        assert result is None

    @pytest.mark.asyncio
    async def test_export_photos_no_reko_reports(
        self, db_session: AsyncSession, export_event: Event, export_incident: Incident
    ):
        """Test photos export with incident but no Reko reports returns None."""
        result = await export_event_photos(db_session, export_event.id)

        assert result is None

    @pytest.mark.asyncio
    async def test_export_photos_draft_reko_excluded(
        self, db_session: AsyncSession, export_event: Event, export_incident: Incident
    ):
        """Test that draft Reko reports are excluded from photo export."""
        import secrets

        # Create draft Reko report with photos
        report = RekoReport(
            id=uuid4(),
            incident_id=export_incident.id,
            token=secrets.token_urlsafe(32),
            is_draft=True,  # Draft - should be excluded
            photos_json=[{"filename": "test_photo.jpg"}],
        )
        db_session.add(report)
        await db_session.commit()

        result = await export_event_photos(db_session, export_event.id)

        # Should return None because draft Reko is excluded
        assert result is None

    @pytest.mark.asyncio
    async def test_export_photos_reko_without_photos(
        self,
        db_session: AsyncSession,
        export_event: Event,
        export_incident: Incident,
        export_reko_report: RekoReport,
    ):
        """Test photos export with Reko report but no photos."""
        # export_reko_report doesn't have photos_json set
        result = await export_event_photos(db_session, export_event.id)

        assert result is None

    @pytest.mark.asyncio
    async def test_export_photos_empty_photos_json(
        self, db_session: AsyncSession, export_event: Event, export_incident: Incident
    ):
        """Test photos export with empty photos_json list."""
        import secrets

        report = RekoReport(
            id=uuid4(),
            incident_id=export_incident.id,
            token=secrets.token_urlsafe(32),
            is_draft=False,
            photos_json=[],  # Empty list
        )
        db_session.add(report)
        await db_session.commit()

        result = await export_event_photos(db_session, export_event.id)

        assert result is None

    @pytest.mark.asyncio
    async def test_export_photos_malformed_photos_json(
        self, db_session: AsyncSession, export_event: Event, export_incident: Incident
    ):
        """Test photos export handles malformed photos_json gracefully."""
        import secrets

        report = RekoReport(
            id=uuid4(),
            incident_id=export_incident.id,
            token=secrets.token_urlsafe(32),
            is_draft=False,
            photos_json=[{"no_filename": "value"}, "invalid_entry"],  # Malformed
        )
        db_session.add(report)
        await db_session.commit()

        # Should not raise, just return None (no valid photos)
        result = await export_event_photos(db_session, export_event.id)

        assert result is None

    @pytest.mark.asyncio
    async def test_export_photos_file_not_exists(
        self, db_session: AsyncSession, export_event: Event, export_incident: Incident
    ):
        """Test photos export when photo file doesn't exist on disk."""
        import secrets

        report = RekoReport(
            id=uuid4(),
            incident_id=export_incident.id,
            token=secrets.token_urlsafe(32),
            is_draft=False,
            photos_json=[{"filename": "nonexistent_photo_12345.jpg"}],
        )
        db_session.add(report)
        await db_session.commit()

        # Should not raise, just return None (file doesn't exist)
        result = await export_event_photos(db_session, export_event.id)

        assert result is None

    @pytest.mark.asyncio
    async def test_export_photos_excludes_deleted_incidents(
        self, db_session: AsyncSession, export_event: Event, export_incident: Incident
    ):
        """Test that photos from soft-deleted incidents are excluded."""
        import secrets
        from datetime import datetime

        # Create Reko with photo reference
        report = RekoReport(
            id=uuid4(),
            incident_id=export_incident.id,
            token=secrets.token_urlsafe(32),
            is_draft=False,
            photos_json=[{"filename": "test_photo.jpg"}],
        )
        db_session.add(report)

        # Soft delete the incident
        export_incident.deleted_at = datetime.now(UTC)
        await db_session.commit()

        result = await export_event_photos(db_session, export_event.id)

        # Should return None because incident is deleted
        assert result is None


# ============================================
# Edge Cases and Boundary Tests
# ============================================


class TestExportEdgeCases:
    """Edge case tests for export functionality."""

    @pytest.mark.asyncio
    async def test_export_incident_with_long_title(
        self, db_session: AsyncSession, export_event: Event, export_user: User
    ):
        """Test export handles long incident titles (truncation in PDF)."""
        long_title = "A" * 200  # Very long title

        incident = Incident(
            id=uuid4(),
            title=long_title,
            type="brandbekaempfung",
            priority="high",
            status="eingegangen",
            event_id=export_event.id,
            created_by=export_user.id,
        )
        db_session.add(incident)
        await db_session.commit()

        # PDF should handle truncation
        pdf_result = await export_event_pdf(db_session, export_event.id, export_user)
        assert pdf_result is not None
        assert pdf_result.read()[:4] == b"%PDF"

        # Excel should store full title
        pdf_result.seek(0)  # Reset for next read
        excel_result = await export_event_excel(db_session, export_event.id)
        wb = load_workbook(excel_result)
        ws_incidents = wb["Einsätze"]
        assert ws_incidents.cell(row=2, column=2).value == long_title

    @pytest.mark.asyncio
    async def test_export_incident_with_special_characters(
        self, db_session: AsyncSession, export_event: Event, export_user: User
    ):
        """Test export handles special characters in text fields."""
        incident = Incident(
            id=uuid4(),
            title="Test <>&'\" Special Chars äöü",
            type="brandbekaempfung",
            priority="high",
            status="eingegangen",
            location_address="Müllerstraße 123 <Basel>",
            description='Description with "quotes" and <tags>',
            event_id=export_event.id,
            created_by=export_user.id,
        )
        db_session.add(incident)
        await db_session.commit()

        # Both exports should handle special chars
        pdf_result = await export_event_pdf(db_session, export_event.id, export_user)
        assert pdf_result is not None

        excel_result = await export_event_excel(db_session, export_event.id)
        assert excel_result is not None

    @pytest.mark.asyncio
    async def test_export_incident_without_optional_fields(
        self, db_session: AsyncSession, export_event: Event, export_user: User
    ):
        """Test export handles incidents with missing optional fields."""
        incident = Incident(
            id=uuid4(),
            title="Minimal Incident",
            type="diverse_einsaetze",
            priority="low",
            status="eingegangen",
            event_id=export_event.id,
            created_by=export_user.id,
            # No location_address, location_lat, location_lng, description
        )
        db_session.add(incident)
        await db_session.commit()

        pdf_result = await export_event_pdf(db_session, export_event.id, export_user)
        assert pdf_result is not None

        excel_result = await export_event_excel(db_session, export_event.id)
        assert excel_result is not None

    @pytest.mark.asyncio
    async def test_export_reko_with_null_relevance(
        self, db_session: AsyncSession, export_event: Event, export_incident: Incident, export_user: User
    ):
        """Test export handles Reko with is_relevant = None."""
        import secrets

        report = RekoReport(
            id=uuid4(),
            incident_id=export_incident.id,
            token=secrets.token_urlsafe(32),
            is_relevant=None,  # Not yet determined
            is_draft=False,
        )
        db_session.add(report)
        await db_session.commit()

        pdf_result = await export_event_pdf(db_session, export_event.id, export_user)
        assert pdf_result is not None

    @pytest.mark.asyncio
    async def test_export_assignment_with_deleted_resource(
        self, db_session: AsyncSession, export_event: Event, export_incident: Incident, export_user: User
    ):
        """Test export handles assignment where resource no longer exists."""
        fake_resource_id = uuid4()

        assignment = IncidentAssignment(
            id=uuid4(),
            incident_id=export_incident.id,
            resource_type="personnel",
            resource_id=fake_resource_id,  # Non-existent
        )
        db_session.add(assignment)
        await db_session.commit()

        # Should not raise, should show "Unbekannt"
        pdf_result = await export_event_pdf(db_session, export_event.id, export_user)
        assert pdf_result is not None

        excel_result = await export_event_excel(db_session, export_event.id)
        assert excel_result is not None
