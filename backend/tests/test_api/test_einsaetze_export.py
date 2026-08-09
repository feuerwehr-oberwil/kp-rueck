"""Tests for the Einsätze export endpoint (plan 25, §7).

``GET /api/exports/events/{event_id}/einsaetze.xlsx`` — editor-only,
rate-limited like its neighbours on the exports router, 404 for unknown events.
"""

import io
from datetime import UTC, datetime
from uuid import uuid4

import openpyxl
import pytest
from httpx import AsyncClient

from app.models import Event, Incident, SchadenplatzReport

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class TestEinsaetzeExportAuth:
    @pytest.mark.asyncio
    async def test_unauthenticated_gets_401(self, client: AsyncClient, test_event: Event):
        response = await client.get(f"/api/exports/events/{test_event.id}/einsaetze.xlsx")
        assert response.status_code == 401

    @pytest.mark.asyncio
    async def test_viewer_gets_403(self, viewer_client: AsyncClient, test_event: Event):
        """Owner names and addresses are citizen PII (§9) — editors only."""
        response = await viewer_client.get(f"/api/exports/events/{test_event.id}/einsaetze.xlsx")
        assert response.status_code == 403

    def test_rate_limited_like_its_neighbours(self):
        """Rate limiting is globally disabled in tests, so assert the registration."""
        import app.main  # noqa: F401  (registers the routers, and with them the limits)
        from app.middleware.rate_limit import RateLimits, limiter

        limits = limiter._route_limits["app.api.exports.export_event_einsaetze"]
        neighbour = limiter._route_limits["app.api.exports.export_event_report"]
        assert [str(item.limit) for item in limits] == [str(item.limit) for item in neighbour]
        assert RateLimits.EXPORT == "10/minute"


class TestEinsaetzeExportSuccess:
    @pytest.mark.asyncio
    async def test_returns_xlsx_with_correct_headers(
        self, editor_client: AsyncClient, test_event: Event, test_incident: Incident
    ):
        response = await editor_client.get(f"/api/exports/events/{test_event.id}/einsaetze.xlsx")
        assert response.status_code == 200
        assert response.headers["content-type"] == XLSX_MEDIA_TYPE

        disposition = response.headers.get("content-disposition", "")
        assert "attachment" in disposition
        assert "einsaetze-test-event-" in disposition
        assert disposition.endswith('.xlsx"')

    @pytest.mark.asyncio
    async def test_schadenplatz_without_a_rapport_still_gets_a_row(
        self, editor_client: AsyncClient, test_event: Event, test_incident: Incident
    ):
        """There is no acceptance step (decision 10): the gaps must be visible."""
        response = await editor_client.get(f"/api/exports/events/{test_event.id}/einsaetze.xlsx")
        assert response.status_code == 200

        sheet = openpyxl.load_workbook(io.BytesIO(response.content))["Einsätze"]
        assert sheet.max_row == 2  # header + the one incident, which has no rapport
        assert sheet.cell(row=2, column=1).value == 1
        assert sheet.cell(row=2, column=2).value == (test_incident.location_address or "")

    @pytest.mark.asyncio
    async def test_empty_event_is_just_the_header(self, editor_client: AsyncClient, test_event: Event):
        response = await editor_client.get(f"/api/exports/events/{test_event.id}/einsaetze.xlsx")
        assert response.status_code == 200
        sheet = openpyxl.load_workbook(io.BytesIO(response.content))["Einsätze"]
        assert sheet.max_row == 1

    @pytest.mark.asyncio
    async def test_a_filed_rapport_reaches_the_sheet(
        self,
        editor_client: AsyncClient,
        db_session,
        test_event: Event,
        test_incident: Incident,
        test_personnel,
    ):
        """End to end through ``collect_event_report_data``, not a hand-built object."""
        test_incident.field_complete_reported_at = datetime(2026, 6, 1, 11, 10, tzinfo=UTC)
        db_session.add(
            SchadenplatzReport(
                incident_id=test_incident.id,
                # No stored times: Beginn/Ende/Dauer are derived from the board.
                arrived_at=datetime(2026, 6, 1, 9, 30, tzinfo=UTC),
                kurzbericht="Baum von der Fahrbahn geräumt.",
                materials_json=[
                    {"assignment_id": str(uuid4()), "name": "Motorsäge", "used": True, "left_on_site": False},
                    {"assignment_id": str(uuid4()), "name": "Beleuchtung", "used": None, "left_on_site": False},
                ],
                is_draft=False,
                submitted_at=datetime(2026, 6, 1, 12, 32, tzinfo=UTC),
                created_by_personnel_id=test_personnel.id,
                updated_by_personnel_id=test_personnel.id,
            )
        )
        await db_session.commit()

        response = await editor_client.get(f"/api/exports/events/{test_event.id}/einsaetze.xlsx")
        assert response.status_code == 200

        sheet = openpyxl.load_workbook(io.BytesIO(response.content))["Einsätze"]
        row = {
            sheet.cell(row=1, column=i).value: sheet.cell(row=2, column=i).value for i in range(1, sheet.max_column + 1)
        }
        assert row["Dauer"] == "1:40"
        assert row["Kurzbericht"] == "Baum von der Fahrbahn geräumt."
        assert "Motorsäge: gebraucht" in row["Material gebraucht"]
        assert "Beleuchtung: keine Angabe" in row["Material gebraucht"]
        assert f"Erfasst von {test_personnel.name} (Feld)" in row["Erfasst von"]

    @pytest.mark.asyncio
    async def test_the_fahrzeuge_column_lists_the_ticked_names(
        self,
        editor_client: AsyncClient,
        db_session,
        test_event: Event,
        test_incident: Incident,
        test_vehicle,
    ):
        """The crew confirms WHICH vehicles; an unticked one is simply not on the line."""
        absent_id = uuid4()
        db_session.add(
            SchadenplatzReport(
                incident_id=test_incident.id,
                vehicles_json=[
                    {
                        "assignment_id": str(uuid4()),
                        "vehicle_id": str(test_vehicle.id),
                        "name": test_vehicle.name,
                        "present": True,
                    },
                    {
                        "assignment_id": str(uuid4()),
                        "vehicle_id": str(absent_id),
                        "name": "MTW",
                        "present": False,
                    },
                ],
                is_draft=False,
                submitted_at=datetime(2026, 6, 1, 12, 32, tzinfo=UTC),
            )
        )
        await db_session.commit()

        response = await editor_client.get(f"/api/exports/events/{test_event.id}/einsaetze.xlsx")
        assert response.status_code == 200

        sheet = openpyxl.load_workbook(io.BytesIO(response.content))["Einsätze"]
        row = {
            sheet.cell(row=1, column=i).value: sheet.cell(row=2, column=i).value for i in range(1, sheet.max_column + 1)
        }
        assert row["Fahrzeuge"] == test_vehicle.name
        assert "MTW" not in str(row["Fahrzeuge"])

    @pytest.mark.asyncio
    async def test_unknown_event_404(self, editor_client: AsyncClient):
        response = await editor_client.get(f"/api/exports/events/{uuid4()}/einsaetze.xlsx")
        assert response.status_code == 404
