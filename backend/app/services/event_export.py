"""Event export service - generates ZIP archives of event data."""
import io
import json
from datetime import datetime
from typing import Any
from uuid import UUID
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from ..models import (
    Event,
    Incident,
    IncidentAssignment,
    RekoReport,
    StatusTransition,
)


def _to_serializable(obj: Any) -> Any:
    """Convert objects to JSON-serializable format."""
    if isinstance(obj, datetime):
        return obj.isoformat()
    if isinstance(obj, UUID):
        return str(obj)
    if isinstance(obj, dict):
        return {k: _to_serializable(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_to_serializable(item) for item in obj]
    return obj


async def export_event_to_zip(db: AsyncSession, event_id: str) -> io.BytesIO:
    """
    Export event and all related data to a ZIP archive.

    Includes:
    - Event metadata (JSON)
    - All incidents (JSON + Excel)
    - Status transitions (JSON)
    - Assignments (JSON)
    - Reko reports (JSON)

    Args:
        db: Database session
        event_id: Event UUID to export

    Returns:
        BytesIO buffer containing ZIP archive
    """
    # Fetch event with all related data
    result = await db.execute(
        select(Event)
        .where(Event.id == UUID(event_id))
        .options(
            selectinload(Event.incidents).selectinload(Incident.assignments),
            selectinload(Event.incidents).selectinload(Incident.status_transitions),
            selectinload(Event.incidents).selectinload(Incident.reko_reports),
        )
    )
    event = result.scalar_one_or_none()

    if not event:
        raise ValueError(f"Event {event_id} not found")

    # Create ZIP buffer
    zip_buffer = io.BytesIO()

    with ZipFile(zip_buffer, 'w', ZIP_DEFLATED) as zip_file:
        # 1. Event metadata
        event_data = {
            "id": str(event.id),
            "name": event.name,
            "training_flag": event.training_flag,
            "created_at": event.created_at.isoformat(),
            "updated_at": event.updated_at.isoformat(),
            "archived_at": event.archived_at.isoformat() if event.archived_at else None,
            "last_activity_at": event.last_activity_at.isoformat(),
            "incident_count": len(event.incidents),
        }
        zip_file.writestr(
            "event_metadata.json",
            json.dumps(event_data, indent=2, ensure_ascii=False)
        )

        # 2. Incidents data
        incidents_data = []
        assignments_data = []
        transitions_data = []
        reko_reports_data = []

        for incident in event.incidents:
            incident_dict = {
                "id": str(incident.id),
                "event_id": str(incident.event_id),
                "title": incident.title,
                "type": incident.type,
                "priority": incident.priority,
                "location_address": incident.location_address,
                "location_lat": str(incident.location_lat) if incident.location_lat else None,
                "location_lng": str(incident.location_lng) if incident.location_lng else None,
                "status": incident.status,
                "description": incident.description,
                "created_at": incident.created_at.isoformat(),
                "updated_at": incident.updated_at.isoformat(),
                "completed_at": incident.completed_at.isoformat() if incident.completed_at else None,
            }
            incidents_data.append(incident_dict)

            # Assignments
            for assignment in incident.assignments:
                assignments_data.append({
                    "id": str(assignment.id),
                    "incident_id": str(assignment.incident_id),
                    "resource_type": assignment.resource_type,
                    "resource_id": str(assignment.resource_id),
                    "assigned_at": assignment.assigned_at.isoformat(),
                    "assigned_by": str(assignment.assigned_by) if assignment.assigned_by else None,
                    "unassigned_at": assignment.unassigned_at.isoformat() if assignment.unassigned_at else None,
                })

            # Status transitions
            for transition in incident.status_transitions:
                transitions_data.append({
                    "id": str(transition.id),
                    "incident_id": str(transition.incident_id),
                    "from_status": transition.from_status,
                    "to_status": transition.to_status,
                    "timestamp": transition.timestamp.isoformat(),
                    "user_id": str(transition.user_id) if transition.user_id else None,
                    "notes": transition.notes,
                })

            # Reko reports
            for report in incident.reko_reports:
                reko_reports_data.append({
                    "id": str(report.id),
                    "incident_id": str(report.incident_id),
                    "submitted_at": report.submitted_at.isoformat(),
                    "updated_at": report.updated_at.isoformat(),
                    "is_relevant": report.is_relevant,
                    "dangers_json": report.dangers_json,
                    "effort_json": report.effort_json,
                    "power_supply": report.power_supply,
                    "photos_json": report.photos_json,
                    "summary_text": report.summary_text,
                    "additional_notes": report.additional_notes,
                    "is_draft": report.is_draft,
                })

        # Write JSON files
        zip_file.writestr(
            "incidents.json",
            json.dumps(incidents_data, indent=2, ensure_ascii=False)
        )
        zip_file.writestr(
            "assignments.json",
            json.dumps(assignments_data, indent=2, ensure_ascii=False)
        )
        zip_file.writestr(
            "status_transitions.json",
            json.dumps(transitions_data, indent=2, ensure_ascii=False)
        )
        zip_file.writestr(
            "reko_reports.json",
            json.dumps(reko_reports_data, indent=2, ensure_ascii=False)
        )

        # 3. Create Excel summary
        excel_buffer = _create_incidents_excel(event, incidents_data)
        zip_file.writestr("incidents_summary.xlsx", excel_buffer.getvalue())

        # 4. Add README
        readme_content = f"""Event Export: {event.name}
====================={"=" * len(event.name)}

Export Date: {datetime.utcnow().isoformat()}
Event ID: {event.id}
Training Mode: {'Yes' if event.training_flag else 'No'}
Created: {event.created_at.isoformat()}
Incident Count: {len(event.incidents)}

Files Included:
- event_metadata.json: Event details
- incidents.json: All incidents ({len(incidents_data)} total)
- assignments.json: Resource assignments ({len(assignments_data)} total)
- status_transitions.json: Status change history ({len(transitions_data)} total)
- reko_reports.json: Field reconnaissance reports ({len(reko_reports_data)} total)
- incidents_summary.xlsx: Excel summary of all incidents
- README.txt: This file

JSON Format:
All JSON files use UTF-8 encoding and ISO 8601 datetime format.
UUIDs are stored as strings.

Excel Summary:
The incidents_summary.xlsx file contains a tabular view of all incidents
for easy viewing in Excel or Google Sheets.
"""
        zip_file.writestr("README.txt", readme_content)

    zip_buffer.seek(0)
    return zip_buffer


def _create_incidents_excel(event: Event, incidents_data: list[dict]) -> io.BytesIO:
    """Create Excel workbook with incidents summary."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Incidents"

    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Headers
    headers = [
        "ID", "Title", "Type", "Priority", "Status",
        "Location", "Description", "Created At", "Completed At"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for row_num, incident in enumerate(incidents_data, 2):
        ws.cell(row=row_num, column=1, value=incident["id"])
        ws.cell(row=row_num, column=2, value=incident["title"])
        ws.cell(row=row_num, column=3, value=incident["type"])
        ws.cell(row=row_num, column=4, value=incident["priority"])
        ws.cell(row=row_num, column=5, value=incident["status"])
        ws.cell(row=row_num, column=6, value=incident["location_address"])
        ws.cell(row=row_num, column=7, value=incident["description"])
        ws.cell(row=row_num, column=8, value=incident["created_at"])
        ws.cell(row=row_num, column=9, value=incident["completed_at"])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer
