"""Event export service for legal compliance and archival."""

import os
import uuid
import zipfile
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ..models import (
    AuditLog,
    Event,
    Incident,
    IncidentAssignment,
    Material,
    Personnel,
    RekoReport,
    StatusTransition,
    User,
    Vehicle,
)


async def export_event_pdf(db: AsyncSession, event_id: uuid.UUID, user: User) -> BytesIO:
    """
    Export event to PDF/A format.

    Structure:
    - Cover page with event metadata
    - Summary table of all incidents
    - Individual incident detail pages
    - Signature page

    Args:
        db: Database session
        event_id: Event ID to export
        user: User performing the export

    Returns:
        BytesIO buffer containing the PDF
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    story = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=24,
        spaceAfter=30,
        alignment=1,  # Center
    )

    heading_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontSize=14,
        spaceAfter=12,
        spaceBefore=12,
    )

    # Get event data
    event_result = await db.execute(select(Event).where(Event.id == event_id))
    event = event_result.scalar_one_or_none()

    if not event:
        raise ValueError(f"Event {event_id} not found")

    # Get all incidents for this event (excluding soft-deleted)
    incidents_result = await db.execute(
        select(Incident)
        .where(Incident.event_id == event_id)
        .where(Incident.deleted_at.is_(None))
        .order_by(Incident.created_at.asc())
    )
    incidents = list(incidents_result.scalars().all())

    # ========== Cover Page ==========
    story.append(Paragraph("Einsatzbericht", title_style))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"<b>Event:</b> {event.name}", styles["Normal"]))
    story.append(Paragraph(f"<b>Datum:</b> {event.created_at.strftime('%d.%m.%Y')}", styles["Normal"]))
    story.append(Paragraph(f"<b>Anzahl Einsätze:</b> {len(incidents)}", styles["Normal"]))
    story.append(Paragraph(f"<b>Trainingsmodus:</b> {'Ja' if event.training_flag else 'Nein'}", styles["Normal"]))
    story.append(Spacer(1, 24))
    story.append(Paragraph(f"<b>Exportiert von:</b> {user.username}", styles["Normal"]))
    story.append(
        Paragraph(f"<b>Exportiert am:</b> {datetime.utcnow().strftime('%d.%m.%Y %H:%M')} UTC", styles["Normal"])
    )
    story.append(PageBreak())

    # ========== Summary Table ==========
    story.append(Paragraph("Übersicht aller Einsätze", heading_style))
    story.append(Spacer(1, 12))

    table_data = [["Nr", "Titel", "Typ", "Priorität", "Status", "Erstellt", "Abgeschlossen"]]

    for i, inc in enumerate(incidents, 1):
        table_data.append(
            [
                str(i),
                inc.title[:30],  # Truncate long titles
                inc.type,
                inc.priority,
                inc.status,
                inc.created_at.strftime("%d.%m %H:%M"),
                inc.completed_at.strftime("%d.%m %H:%M") if inc.completed_at else "-",
            ]
        )

    table = Table(table_data, colWidths=[1 * cm, 5 * cm, 3 * cm, 2 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
            ]
        )
    )
    story.append(table)
    story.append(PageBreak())

    # ========== Individual Incident Pages ==========
    for idx, incident in enumerate(incidents, 1):
        story.append(Paragraph(f"Einsatz {idx}: {incident.title}", heading_style))
        story.append(Spacer(1, 6))

        # Basic Info
        story.append(Paragraph("<b>Grundinformationen</b>", styles["Heading3"]))
        story.append(Paragraph(f"Typ: {incident.type}", styles["Normal"]))
        story.append(Paragraph(f"Priorität: {incident.priority}", styles["Normal"]))
        if incident.location_address:
            story.append(Paragraph(f"Adresse: {incident.location_address}", styles["Normal"]))
        if incident.description:
            story.append(Paragraph(f"Beschreibung: {incident.description}", styles["Normal"]))
        story.append(Spacer(1, 12))

        # Timeline (Status Transitions)
        story.append(Paragraph("<b>Zeitverlauf</b>", styles["Heading3"]))
        transitions_result = await db.execute(
            select(StatusTransition)
            .where(StatusTransition.incident_id == incident.id)
            .order_by(StatusTransition.timestamp.asc())
        )
        transitions = list(transitions_result.scalars().all())

        if transitions:
            for transition in transitions:
                story.append(
                    Paragraph(
                        f"{transition.timestamp.strftime('%d.%m.%Y %H:%M')} - "
                        f"{transition.from_status} → {transition.to_status}",
                        styles["Normal"],
                    )
                )
        else:
            story.append(Paragraph(f"Erstellt: {incident.created_at.strftime('%d.%m.%Y %H:%M')}", styles["Normal"]))
        story.append(Spacer(1, 12))

        # Assignments (Personnel, Vehicles, Materials)
        story.append(Paragraph("<b>Zugewiesene Ressourcen</b>", styles["Heading3"]))
        assignments_result = await db.execute(
            select(IncidentAssignment)
            .where(IncidentAssignment.incident_id == incident.id)
            .where(IncidentAssignment.unassigned_at.is_(None))
            .order_by(IncidentAssignment.assigned_at.asc())
        )
        assignments = list(assignments_result.scalars().all())

        if assignments:
            for assignment in assignments:
                # Get resource details based on type
                resource_name = "Unbekannt"
                if assignment.resource_type == "personnel":
                    res = await db.execute(select(Personnel).where(Personnel.id == assignment.resource_id))
                    person = res.scalar_one_or_none()
                    if person:
                        resource_name = f"{person.name} ({person.role or 'N/A'})"
                elif assignment.resource_type == "vehicle":
                    res = await db.execute(select(Vehicle).where(Vehicle.id == assignment.resource_id))
                    vehicle = res.scalar_one_or_none()
                    if vehicle:
                        resource_name = f"{vehicle.name} ({vehicle.type})"
                elif assignment.resource_type == "material":
                    res = await db.execute(select(Material).where(Material.id == assignment.resource_id))
                    material = res.scalar_one_or_none()
                    if material:
                        resource_name = f"{material.name}"

                story.append(
                    Paragraph(
                        f"{assignment.resource_type.capitalize()}: {resource_name} "
                        f"(ab {assignment.assigned_at.strftime('%H:%M')})",
                        styles["Normal"],
                    )
                )
        else:
            story.append(Paragraph("Keine Ressourcen zugewiesen", styles["Normal"]))
        story.append(Spacer(1, 12))

        # Reko Report
        reko_result = await db.execute(
            select(RekoReport)
            .where(RekoReport.incident_id == incident.id)
            .where(not RekoReport.is_draft)
            .order_by(RekoReport.submitted_at.desc())
            .limit(1)
        )
        reko = reko_result.scalar_one_or_none()

        if reko:
            story.append(Paragraph("<b>Reko-Bericht</b>", styles["Heading3"]))
            if reko.is_relevant is not None:
                story.append(
                    Paragraph(f"Relevanz: {'Relevant' if reko.is_relevant else 'Nicht relevant'}", styles["Normal"])
                )
            if reko.summary_text:
                story.append(Paragraph(f"Zusammenfassung: {reko.summary_text}", styles["Normal"]))
            if reko.additional_notes:
                story.append(Paragraph(f"Zusätzliche Notizen: {reko.additional_notes}", styles["Normal"]))
            story.append(Spacer(1, 12))

        # Audit Log
        story.append(Paragraph("<b>Änderungsprotokoll</b>", styles["Heading3"]))
        audit_result = await db.execute(
            select(AuditLog, User)
            .outerjoin(User, AuditLog.user_id == User.id)
            .where(AuditLog.resource_type == "incident")
            .where(AuditLog.resource_id == incident.id)
            .order_by(AuditLog.timestamp.asc())
        )
        audit_entries = list(audit_result.all())

        if audit_entries:
            for audit, audit_user in audit_entries:
                username = audit_user.username if audit_user else "System"
                story.append(
                    Paragraph(
                        f"{audit.timestamp.strftime('%d.%m.%Y %H:%M')} - {username}: {audit.action_type}",
                        styles["Normal"],
                    )
                )
        else:
            story.append(Paragraph("Keine Protokolleinträge vorhanden", styles["Normal"]))

        # Add page break between incidents (except for last one)
        if idx < len(incidents):
            story.append(PageBreak())

    # ========== Signature Page ==========
    story.append(PageBreak())
    story.append(Paragraph("Unterschriften", heading_style))
    story.append(Spacer(1, 48))
    story.append(Paragraph("_" * 60, styles["Normal"]))
    story.append(Paragraph("Einsatzleiter", styles["Normal"]))
    story.append(Spacer(1, 24))
    story.append(Paragraph("_" * 60, styles["Normal"]))
    story.append(Paragraph("Datum", styles["Normal"]))

    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


async def export_event_excel(db: AsyncSession, event_id: uuid.UUID) -> BytesIO:
    """
    Export event to Excel format (same data as PDF).

    Args:
        db: Database session
        event_id: Event ID to export

    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Get event data
    event_result = await db.execute(select(Event).where(Event.id == event_id))
    event = event_result.scalar_one_or_none()

    if not event:
        raise ValueError(f"Event {event_id} not found")

    # Get all incidents
    incidents_result = await db.execute(
        select(Incident)
        .where(Incident.event_id == event_id)
        .where(Incident.deleted_at.is_(None))
        .order_by(Incident.created_at.asc())
    )
    incidents = list(incidents_result.scalars().all())

    # ========== Overview Sheet ==========
    ws_overview = wb.create_sheet("Übersicht")
    ws_overview["A1"] = "Einsatzbericht"
    ws_overview["A1"].font = Font(size=16, bold=True)

    ws_overview["A3"] = "Event:"
    ws_overview["B3"] = event.name
    ws_overview["A4"] = "Datum:"
    ws_overview["B4"] = event.created_at.strftime("%d.%m.%Y")
    ws_overview["A5"] = "Anzahl Einsätze:"
    ws_overview["B5"] = len(incidents)
    ws_overview["A6"] = "Trainingsmodus:"
    ws_overview["B6"] = "Ja" if event.training_flag else "Nein"

    # ========== Incidents Summary Sheet ==========
    ws_incidents = wb.create_sheet("Einsätze")

    # Headers
    headers = ["Nr", "Titel", "Typ", "Priorität", "Status", "Adresse", "Erstellt", "Abgeschlossen"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_incidents.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    # Data rows
    for row_num, (idx, incident) in enumerate(enumerate(incidents, 1), 2):
        ws_incidents.cell(row=row_num, column=1, value=idx)
        ws_incidents.cell(row=row_num, column=2, value=incident.title)
        ws_incidents.cell(row=row_num, column=3, value=incident.type)
        ws_incidents.cell(row=row_num, column=4, value=incident.priority)
        ws_incidents.cell(row=row_num, column=5, value=incident.status)
        ws_incidents.cell(row=row_num, column=6, value=incident.location_address or "")
        ws_incidents.cell(row=row_num, column=7, value=incident.created_at.strftime("%d.%m.%Y %H:%M"))
        ws_incidents.cell(
            row=row_num,
            column=8,
            value=incident.completed_at.strftime("%d.%m.%Y %H:%M") if incident.completed_at else "",
        )

    # ========== Detailed Incidents Sheet ==========
    ws_details = wb.create_sheet("Einsatz Details")
    current_row = 1

    for idx, incident in enumerate(incidents, 1):
        # Incident header
        ws_details.cell(row=current_row, column=1, value=f"Einsatz {idx}: {incident.title}")
        ws_details.cell(row=current_row, column=1).font = Font(size=14, bold=True)
        current_row += 2

        # Basic info
        ws_details.cell(row=current_row, column=1, value="Typ:")
        ws_details.cell(row=current_row, column=2, value=incident.type)
        current_row += 1
        ws_details.cell(row=current_row, column=1, value="Priorität:")
        ws_details.cell(row=current_row, column=2, value=incident.priority)
        current_row += 1
        if incident.location_address:
            ws_details.cell(row=current_row, column=1, value="Adresse:")
            ws_details.cell(row=current_row, column=2, value=incident.location_address)
            current_row += 1
        if incident.description:
            ws_details.cell(row=current_row, column=1, value="Beschreibung:")
            ws_details.cell(row=current_row, column=2, value=incident.description)
            current_row += 1
        current_row += 1

        # Status transitions
        ws_details.cell(row=current_row, column=1, value="Zeitverlauf:")
        ws_details.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

        transitions_result = await db.execute(
            select(StatusTransition)
            .where(StatusTransition.incident_id == incident.id)
            .order_by(StatusTransition.timestamp.asc())
        )
        transitions = list(transitions_result.scalars().all())

        for transition in transitions:
            ws_details.cell(
                row=current_row,
                column=1,
                value=f"{transition.timestamp.strftime('%d.%m.%Y %H:%M')} - "
                f"{transition.from_status} → {transition.to_status}",
            )
            current_row += 1
        current_row += 1

        # Assignments
        ws_details.cell(row=current_row, column=1, value="Zugewiesene Ressourcen:")
        ws_details.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

        assignments_result = await db.execute(
            select(IncidentAssignment)
            .where(IncidentAssignment.incident_id == incident.id)
            .where(IncidentAssignment.unassigned_at.is_(None))
        )
        assignments = list(assignments_result.scalars().all())

        for assignment in assignments:
            resource_name = "Unbekannt"
            if assignment.resource_type == "personnel":
                res = await db.execute(select(Personnel).where(Personnel.id == assignment.resource_id))
                person = res.scalar_one_or_none()
                if person:
                    resource_name = f"{person.name}"
            elif assignment.resource_type == "vehicle":
                res = await db.execute(select(Vehicle).where(Vehicle.id == assignment.resource_id))
                vehicle = res.scalar_one_or_none()
                if vehicle:
                    resource_name = f"{vehicle.name}"
            elif assignment.resource_type == "material":
                res = await db.execute(select(Material).where(Material.id == assignment.resource_id))
                material = res.scalar_one_or_none()
                if material:
                    resource_name = f"{material.name}"

            ws_details.cell(
                row=current_row, column=1, value=f"{assignment.resource_type.capitalize()}: {resource_name}"
            )
            current_row += 1

        current_row += 2  # Space between incidents

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


async def export_event_photos(db: AsyncSession, event_id: uuid.UUID) -> BytesIO | None:
    """
    Create ZIP file with all photos from event incidents.

    Args:
        db: Database session
        event_id: Event ID to export

    Returns:
        BytesIO buffer containing the photos ZIP, or None if no photos exist
    """
    buffer = BytesIO()

    # Get all incidents for this event
    incidents_result = await db.execute(
        select(Incident).where(Incident.event_id == event_id).where(Incident.deleted_at.is_(None))
    )
    incidents = list(incidents_result.scalars().all())

    photo_count = 0

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for incident in incidents:
            # Get reko report for incident
            reko_result = await db.execute(
                select(RekoReport).where(RekoReport.incident_id == incident.id).where(not RekoReport.is_draft)
            )
            reko_reports = list(reko_result.scalars().all())

            for reko in reko_reports:
                if reko.photos_json:
                    # photos_json is expected to be a list of photo objects
                    for photo in reko.photos_json:
                        if isinstance(photo, dict) and "filename" in photo:
                            filename = photo["filename"]
                            # Assume photos are stored in a photos/ directory
                            photo_path = f"photos/{filename}"

                            if os.path.exists(photo_path):
                                # Create incident folder in ZIP
                                arcname = f"{incident.title}/{filename}"
                                zf.write(photo_path, arcname=arcname)
                                photo_count += 1

    if photo_count == 0:
        return None  # No photos found

    buffer.seek(0)
    return buffer
