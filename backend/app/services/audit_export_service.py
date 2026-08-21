"""Event audit export service for payment processing."""

import asyncio
import re
import uuid
from dataclasses import dataclass, field
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ..models import (
    AuditLog,
    Event,
    EventAttendance,
    Incident,
    IncidentAssignment,
    Material,
    Personnel,
    RekoReport,
    SchadenplatzReport,
    StatusTransition,
    User,
    Vehicle,
)

# Audit actions that are journal-worthy for the Einsatztagebuch chapter in the
# PDF report. Deliberately a whitelist: field-level updates, logins, exports,
# settings changes etc. must never leak into the after-action journal.
JOURNAL_AUDIT_ACTIONS: tuple[str, ...] = ("divera_alarm", "delete", "restore")


@dataclass
class EventReportData:
    """Fully-loaded event data shared by the Excel and PDF report builders.

    Populated by :func:`collect_event_report_data`. Holds the event, its
    (non-deleted) incidents ordered by ``created_at``, the full assignment
    history, status transitions, reko reports, plus lookup maps keyed by id so
    a consumer can resolve incident/resource/user references without extra
    queries.
    """

    event: Event
    incidents: list[Incident]
    assignments: list[IncidentAssignment]
    transitions: list[StatusTransition]
    reko_reports: list[RekoReport]
    # Journal-worthy audit rows (see JOURNAL_AUDIT_ACTIONS), incident-scoped.
    audit_entries: list[AuditLog] = field(default_factory=list)
    # Schadenplatz-Rapporte (plan 25), at most one per incident. Drafts included:
    # a half-filled rapport is still what the crew said, and the outputs mark it.
    schadenplatz_reports: list[SchadenplatzReport] = field(default_factory=list)
    # Anwesenheit: one row per person who actually arrived, in arrival order.
    # Rows whose `checked_in_at` is NULL are dropped by the collector – a bulk
    # check-out can create an attendance row for somebody who never came, and a
    # departure without an arrival is not evidence of presence.
    attendance: list[EventAttendance] = field(default_factory=list)
    incident_map: dict[uuid.UUID, Incident] = field(default_factory=dict)
    personnel_map: dict[uuid.UUID, Personnel] = field(default_factory=dict)
    vehicle_map: dict[uuid.UUID, Vehicle] = field(default_factory=dict)
    material_map: dict[uuid.UUID, Material] = field(default_factory=dict)
    user_map: dict[uuid.UUID, User] = field(default_factory=dict)


def format_timestamp(dt: datetime | None) -> str:
    """Format datetime to ISO 8601 with timezone."""
    if dt is None:
        return ""
    if dt.tzinfo is None:
        return dt.replace(tzinfo=UTC).isoformat()
    return dt.isoformat()


async def collect_event_report_data(db: AsyncSession, event_id: uuid.UUID) -> EventReportData:
    """Load an event and all data needed to render a report (Excel or PDF).

    This is the single source of truth for the report queries. Both
    :func:`export_event_audit_excel` and the PDF report builder consume the
    returned :class:`EventReportData` so the two exports never drift apart.

    Args:
        db: Database session
        event_id: Event ID to load

    Returns:
        A populated :class:`EventReportData`.

    Raises:
        ValueError: If the event does not exist.
    """
    # ========== 1. Load event ==========
    event_result = await db.execute(select(Event).where(Event.id == event_id))
    event = event_result.scalar_one_or_none()

    if not event:
        raise ValueError(f"Event {event_id} not found")

    # ========== 2. Load incidents (excluding soft-deleted) ==========
    incidents_result = await db.execute(
        select(Incident)
        .where(Incident.event_id == event_id)
        .where(Incident.deleted_at.is_(None))
        .order_by(Incident.created_at.asc())
    )
    incidents = list(incidents_result.scalars().all())

    # ========== 3. Collect all incident IDs ==========
    incident_ids = [inc.id for inc in incidents]
    incident_map = {inc.id: inc for inc in incidents}

    # ========== 4. Batch load ALL assignments (not just active) ==========
    if incident_ids:
        assignments_result = await db.execute(
            select(IncidentAssignment)
            .where(IncidentAssignment.incident_id.in_(incident_ids))
            .order_by(IncidentAssignment.assigned_at)
        )
        assignments = list(assignments_result.scalars().all())
    else:
        assignments = []

    # ========== 5. Extract unique resource IDs and batch load ==========
    personnel_ids = {a.resource_id for a in assignments if a.resource_type == "personnel"}
    vehicle_ids = {a.resource_id for a in assignments if a.resource_type == "vehicle"}
    material_ids = {a.resource_id for a in assignments if a.resource_type == "material"}
    user_ids = {a.assigned_by for a in assignments if a.assigned_by}

    # Also get user IDs from incidents (created_by)
    user_ids.update(inc.created_by for inc in incidents if inc.created_by)

    # Load status transitions
    if incident_ids:
        transitions_result = await db.execute(
            select(StatusTransition)
            .where(StatusTransition.incident_id.in_(incident_ids))
            .order_by(StatusTransition.timestamp.asc())
        )
        transitions = list(transitions_result.scalars().all())
        user_ids.update(t.user_id for t in transitions if t.user_id)
    else:
        transitions = []

    # Load reko reports
    if incident_ids:
        reko_result = await db.execute(
            select(RekoReport).where(RekoReport.incident_id.in_(incident_ids)).order_by(RekoReport.submitted_at.asc())
        )
        reko_reports = list(reko_result.scalars().all())
        # Get personnel IDs from reko reports
        reko_personnel_ids = {r.submitted_by_personnel_id for r in reko_reports if r.submitted_by_personnel_id}
        personnel_ids.update(reko_personnel_ids)
        # …and the KP side of the same provenance (plan 26 §5.3): the operator who
        # typed a dictated report, amended a crew's, or logged "vor Ort" off the
        # radio. Without these the outputs would resolve every one of them to an
        # empty name and print "unbekannt" for people the database knows.
        for reko in reko_reports:
            user_ids.update(
                uid
                for uid in (
                    reko.created_by_user_id,
                    reko.updated_by_user_id,
                    reko.arrived_reported_by_user_id,
                )
                if uid
            )
    else:
        reko_reports = []

    # Load Schadenplatz-Rapporte (plan 25). Their provenance is a pair of FKs —
    # personnel for a `/feld` filing, user for a KP radio entry — and exactly one
    # side is populated per write, so both id sets feed the lookup maps below.
    schadenplatz_reports: list[SchadenplatzReport] = []
    if incident_ids:
        rapport_result = await db.execute(
            select(SchadenplatzReport)
            .where(SchadenplatzReport.incident_id.in_(incident_ids))
            .order_by(SchadenplatzReport.created_at.asc())
        )
        schadenplatz_reports = list(rapport_result.scalars().all())
        for report in schadenplatz_reports:
            personnel_ids.update(pid for pid in (report.created_by_personnel_id, report.updated_by_personnel_id) if pid)
            user_ids.update(uid for uid in (report.created_by_user_id, report.updated_by_user_id) if uid)

    # Load the Anwesenheit (event-scoped, independent of any incident): everybody who
    # actually arrived. `checked_in_at IS NOT NULL` is the filter, not `checked_in` —
    # somebody who has already gone home is exactly who this section is about, while a
    # row created by a check-out for a person who never came carries no arrival and is
    # not attendance. Ordered like every other roll-call in the app (rank, then name:
    # `crud.personnel.get_personnel_list`, the A4 status slip) so the section reads the
    # way the station's printed lists already do.
    attendance_result = await db.execute(
        select(EventAttendance)
        .join(Personnel, EventAttendance.personnel_id == Personnel.id)
        .where(EventAttendance.event_id == event_id)
        .where(EventAttendance.checked_in_at.is_not(None))
        .order_by(Personnel.role_sort_order.asc(), Personnel.role.asc(), Personnel.name.asc())
    )
    attendance = list(attendance_result.scalars().all())
    personnel_ids.update(a.personnel_id for a in attendance)

    # Load journal-worthy audit entries (Einsatztagebuch). Audit rows carry no
    # event scoping, only resource_type/resource_id — so we scope them via the
    # event's incident ids and restrict to the whitelisted action types.
    audit_entries: list[AuditLog] = []
    if incident_ids:
        audit_result = await db.execute(
            select(AuditLog)
            .where(AuditLog.resource_type == "incident")
            .where(AuditLog.resource_id.in_(incident_ids))
            .where(AuditLog.action_type.in_(JOURNAL_AUDIT_ACTIONS))
            .order_by(AuditLog.timestamp.asc())
        )
        audit_entries = list(audit_result.scalars().all())
        user_ids.update(a.user_id for a in audit_entries if a.user_id)

    # ========== 6. Batch load resources and users ==========
    personnel_map: dict[uuid.UUID, Personnel] = {}
    if personnel_ids:
        personnel_result = await db.execute(select(Personnel).where(Personnel.id.in_(personnel_ids)))
        personnel_map = {p.id: p for p in personnel_result.scalars().all()}

    vehicle_map: dict[uuid.UUID, Vehicle] = {}
    if vehicle_ids:
        vehicle_result = await db.execute(select(Vehicle).where(Vehicle.id.in_(vehicle_ids)))
        vehicle_map = {v.id: v for v in vehicle_result.scalars().all()}

    material_map: dict[uuid.UUID, Material] = {}
    if material_ids:
        material_result = await db.execute(select(Material).where(Material.id.in_(material_ids)))
        material_map = {m.id: m for m in material_result.scalars().all()}

    user_map: dict[uuid.UUID, User] = {}
    if user_ids:
        user_result = await db.execute(select(User).where(User.id.in_(user_ids)))
        user_map = {u.id: u for u in user_result.scalars().all()}

    return EventReportData(
        event=event,
        incidents=incidents,
        assignments=assignments,
        transitions=transitions,
        reko_reports=reko_reports,
        audit_entries=audit_entries,
        schadenplatz_reports=schadenplatz_reports,
        attendance=attendance,
        incident_map=incident_map,
        personnel_map=personnel_map,
        vehicle_map=vehicle_map,
        material_map=material_map,
        user_map=user_map,
    )


async def export_event_audit_excel(
    db: AsyncSession, event_id: uuid.UUID, current_user: User
) -> tuple[BytesIO, dict[str, Any]]:
    """
    Export complete event audit data for payment processing.

    Unlike the regular export which only shows currently-assigned resources,
    this export includes the full assignment history with timestamps showing
    when each resource was assigned and released.

    Args:
        db: Database session
        event_id: Event ID to export
        current_user: User performing the export

    Returns:
        Tuple of (BytesIO buffer containing the Excel file, metadata dict)

    Raises:
        ValueError: If event not found
    """
    # Shared data collection (queries live in collect_event_report_data).
    data = await collect_event_report_data(db, event_id)

    # openpyxl sheet building for a large event is pure CPU — keep it off
    # the event loop so the export doesn't freeze every operator's requests
    # and WebSocket pings (audit H4).
    return await asyncio.to_thread(_build_audit_workbook, data, current_user)


def _build_audit_workbook(data: EventReportData, current_user: User) -> tuple[BytesIO, dict[str, Any]]:
    """Blocking workbook construction — runs in a worker thread."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    event = data.event
    incidents = data.incidents
    assignments = data.assignments
    transitions = data.transitions
    reko_reports = data.reko_reports
    incident_map = data.incident_map
    personnel_map = data.personnel_map
    vehicle_map = data.vehicle_map
    material_map = data.material_map
    user_map = data.user_map

    # ========== Sheet 1: Event Overview ==========
    ws_overview = wb.create_sheet("Ereignis Übersicht")
    _add_overview_sheet(ws_overview, event, incidents, assignments, transitions, current_user)

    # ========== Sheet 2: Incidents ==========
    ws_incidents = wb.create_sheet("Einsätze")
    _add_incidents_sheet(ws_incidents, incidents, user_map)

    # ========== Sheet 3: Personnel Assignments ==========
    ws_personnel = wb.create_sheet("Personal Zuweisungen")
    personnel_assignments = [a for a in assignments if a.resource_type == "personnel"]
    _add_personnel_assignments_sheet(ws_personnel, personnel_assignments, incident_map, personnel_map, user_map)

    # ========== Sheet 4: Vehicle Assignments ==========
    ws_vehicles = wb.create_sheet("Fahrzeug Zuweisungen")
    vehicle_assignments = [a for a in assignments if a.resource_type == "vehicle"]
    _add_vehicle_assignments_sheet(ws_vehicles, vehicle_assignments, incident_map, vehicle_map, user_map)

    # ========== Sheet 5: Material Assignments ==========
    ws_materials = wb.create_sheet("Material Zuweisungen")
    material_assignments = [a for a in assignments if a.resource_type == "material"]
    _add_material_assignments_sheet(ws_materials, material_assignments, incident_map, material_map, user_map)

    # ========== Sheet 6: Status Transitions ==========
    ws_transitions = wb.create_sheet("Status Verlauf")
    _add_status_transitions_sheet(ws_transitions, transitions, incident_map, user_map)

    # ========== Sheet 7: Reko Reports ==========
    ws_reko = wb.create_sheet("Reko Berichte")
    _add_reko_reports_sheet(ws_reko, reko_reports, incident_map, personnel_map)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Return metadata for audit logging
    metadata = {
        "exported_at": datetime.now(UTC).isoformat(),
        "incident_count": len(incidents),
        "assignment_count": len(assignments),
        "transition_count": len(transitions),
        "reko_count": len(reko_reports),
    }

    return buffer, metadata


def _add_overview_sheet(
    ws: Worksheet,
    event: Event,
    incidents: list[Incident],
    assignments: list[IncidentAssignment],
    transitions: list[StatusTransition],
    current_user: User,
) -> None:
    """Add event overview sheet."""
    ws["A1"] = "Audit Export - Ereignis Übersicht"
    ws["A1"].font = Font(size=16, bold=True)

    ws["A3"] = "Ereignis Name:"
    ws["B3"] = event.name
    ws["A4"] = "Ereignis ID:"
    ws["B4"] = str(event.id)
    ws["A5"] = "Trainingsmodus:"
    ws["B5"] = "Ja" if event.training_flag else "Nein"
    ws["A6"] = "Erstellt:"
    ws["B6"] = format_timestamp(event.created_at)
    ws["A7"] = "Aktualisiert:"
    ws["B7"] = format_timestamp(event.updated_at)
    ws["A8"] = "Archiviert:"
    ws["B8"] = format_timestamp(event.archived_at) if event.archived_at else "Nein"

    ws["A10"] = "Zusammenfassung"
    ws["A10"].font = Font(size=14, bold=True)

    ws["A11"] = "Anzahl Einsätze:"
    ws["B11"] = len(incidents)
    ws["A12"] = "Anzahl Zuweisungen (gesamt):"
    ws["B12"] = len(assignments)
    ws["A13"] = "Anzahl Status-Übergänge:"
    ws["B13"] = len(transitions)

    ws["A15"] = "Export Metadaten"
    ws["A15"].font = Font(size=14, bold=True)

    ws["A16"] = "Exportiert von:"
    ws["B16"] = current_user.username
    ws["A17"] = "Exportiert am:"
    ws["B17"] = format_timestamp(datetime.now(UTC))

    # Auto-adjust column width
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50


def _add_incidents_sheet(ws: Worksheet, incidents: list[Incident], user_map: dict[uuid.UUID, User]) -> None:
    """Add incidents sheet."""
    headers = [
        "Einsatz ID",
        "Nr",
        "Titel",
        "Typ",
        "Priorität",
        "Status",
        "Nachbarhilfe",
        "Adresse",
        "Lat",
        "Lng",
        "Beschreibung",
        "Kontakt",
        "Erstellt",
        "Erstellt von",
        "Aktualisiert",
        "Abgeschlossen",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    for row_num, (idx, incident) in enumerate(enumerate(incidents, 1), 2):
        ws.cell(row=row_num, column=1, value=str(incident.id))
        ws.cell(row=row_num, column=2, value=idx)
        ws.cell(row=row_num, column=3, value=incident.title)
        ws.cell(row=row_num, column=4, value=incident.type)
        ws.cell(row=row_num, column=5, value=incident.priority)
        ws.cell(row=row_num, column=6, value=incident.status)
        ws.cell(row=row_num, column=7, value="Ja" if incident.nachbarhilfe else "Nein")
        ws.cell(row=row_num, column=8, value=incident.location_address or "")
        ws.cell(row=row_num, column=9, value=str(incident.location_lat) if incident.location_lat else "")
        ws.cell(row=row_num, column=10, value=str(incident.location_lng) if incident.location_lng else "")
        ws.cell(row=row_num, column=11, value=incident.description or "")
        ws.cell(row=row_num, column=12, value=incident.contact or "")
        ws.cell(row=row_num, column=13, value=format_timestamp(incident.created_at))

        # Get creator username
        creator_name = ""
        if incident.created_by and incident.created_by in user_map:
            creator_name = user_map[incident.created_by].username
        ws.cell(row=row_num, column=14, value=creator_name)

        ws.cell(row=row_num, column=15, value=format_timestamp(incident.updated_at))
        ws.cell(row=row_num, column=16, value=format_timestamp(incident.completed_at))


def _add_personnel_assignments_sheet(
    ws: Worksheet,
    assignments: list[IncidentAssignment],
    incident_map: dict[uuid.UUID, Incident],
    personnel_map: dict[uuid.UUID, Personnel],
    user_map: dict[uuid.UUID, User],
) -> None:
    """Add personnel assignments sheet."""
    headers = [
        "Zuweisung ID",
        "Einsatz ID",
        "Einsatz Titel",
        "Personal ID",
        "Personal Name",
        "Rolle",
        "Zugewiesen Um",
        "Zugewiesen Von",
        "Freigegeben Um",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    for row_num, assignment in enumerate(assignments, 2):
        incident = incident_map.get(assignment.incident_id)
        personnel = personnel_map.get(assignment.resource_id)
        assigner = user_map.get(assignment.assigned_by) if assignment.assigned_by else None

        ws.cell(row=row_num, column=1, value=str(assignment.id))
        ws.cell(row=row_num, column=2, value=str(assignment.incident_id))
        ws.cell(row=row_num, column=3, value=incident.title if incident else "")
        ws.cell(row=row_num, column=4, value=str(assignment.resource_id))
        ws.cell(row=row_num, column=5, value=personnel.name if personnel else "")
        ws.cell(row=row_num, column=6, value=personnel.role if personnel else "")
        ws.cell(row=row_num, column=7, value=format_timestamp(assignment.assigned_at))
        ws.cell(row=row_num, column=8, value=assigner.username if assigner else "")
        ws.cell(row=row_num, column=9, value=format_timestamp(assignment.unassigned_at))


def _add_vehicle_assignments_sheet(
    ws: Worksheet,
    assignments: list[IncidentAssignment],
    incident_map: dict[uuid.UUID, Incident],
    vehicle_map: dict[uuid.UUID, Vehicle],
    user_map: dict[uuid.UUID, User],
) -> None:
    """Add vehicle assignments sheet."""
    headers = [
        "Zuweisung ID",
        "Einsatz ID",
        "Einsatz Titel",
        "Fahrzeug ID",
        "Fahrzeug Name",
        "Fahrzeug Typ",
        "Funkrufname",
        "Zugewiesen Um",
        "Zugewiesen Von",
        "Freigegeben Um",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    for row_num, assignment in enumerate(assignments, 2):
        incident = incident_map.get(assignment.incident_id)
        vehicle = vehicle_map.get(assignment.resource_id)
        assigner = user_map.get(assignment.assigned_by) if assignment.assigned_by else None

        ws.cell(row=row_num, column=1, value=str(assignment.id))
        ws.cell(row=row_num, column=2, value=str(assignment.incident_id))
        ws.cell(row=row_num, column=3, value=incident.title if incident else "")
        ws.cell(row=row_num, column=4, value=str(assignment.resource_id))
        ws.cell(row=row_num, column=5, value=vehicle.name if vehicle else "")
        ws.cell(row=row_num, column=6, value=vehicle.type if vehicle else "")
        ws.cell(row=row_num, column=7, value=vehicle.radio_call_sign if vehicle else "")
        ws.cell(row=row_num, column=8, value=format_timestamp(assignment.assigned_at))
        ws.cell(row=row_num, column=9, value=assigner.username if assigner else "")
        ws.cell(row=row_num, column=10, value=format_timestamp(assignment.unassigned_at))


def _add_material_assignments_sheet(
    ws: Worksheet,
    assignments: list[IncidentAssignment],
    incident_map: dict[uuid.UUID, Incident],
    material_map: dict[uuid.UUID, Material],
    user_map: dict[uuid.UUID, User],
) -> None:
    """Add material assignments sheet."""
    headers = [
        "Zuweisung ID",
        "Einsatz ID",
        "Einsatz Titel",
        "Material ID",
        "Material Name",
        "Material Typ",
        "Standort",
        "Zugewiesen Um",
        "Zugewiesen Von",
        "Freigegeben Um",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    for row_num, assignment in enumerate(assignments, 2):
        incident = incident_map.get(assignment.incident_id)
        material = material_map.get(assignment.resource_id)
        assigner = user_map.get(assignment.assigned_by) if assignment.assigned_by else None

        ws.cell(row=row_num, column=1, value=str(assignment.id))
        ws.cell(row=row_num, column=2, value=str(assignment.incident_id))
        ws.cell(row=row_num, column=3, value=incident.title if incident else "")
        ws.cell(row=row_num, column=4, value=str(assignment.resource_id))
        ws.cell(row=row_num, column=5, value=material.name if material else "")
        ws.cell(row=row_num, column=6, value=material.type if material else "")
        ws.cell(row=row_num, column=7, value=material.location if material else "")
        ws.cell(row=row_num, column=8, value=format_timestamp(assignment.assigned_at))
        ws.cell(row=row_num, column=9, value=assigner.username if assigner else "")
        ws.cell(row=row_num, column=10, value=format_timestamp(assignment.unassigned_at))


def _add_status_transitions_sheet(
    ws: Worksheet,
    transitions: list[StatusTransition],
    incident_map: dict[uuid.UUID, Incident],
    user_map: dict[uuid.UUID, User],
) -> None:
    """Add status transitions sheet."""
    headers = [
        "Transition ID",
        "Einsatz ID",
        "Einsatz Titel",
        "Von Status",
        "Zu Status",
        "Zeitstempel",
        "Benutzer",
        "Notizen",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    for row_num, transition in enumerate(transitions, 2):
        incident = incident_map.get(transition.incident_id)
        user = user_map.get(transition.user_id) if transition.user_id else None

        ws.cell(row=row_num, column=1, value=str(transition.id))
        ws.cell(row=row_num, column=2, value=str(transition.incident_id))
        ws.cell(row=row_num, column=3, value=incident.title if incident else "")
        ws.cell(row=row_num, column=4, value=transition.from_status)
        ws.cell(row=row_num, column=5, value=transition.to_status)
        ws.cell(row=row_num, column=6, value=format_timestamp(transition.timestamp))
        ws.cell(row=row_num, column=7, value=user.username if user else "")
        ws.cell(row=row_num, column=8, value=transition.notes or "")


def _add_reko_reports_sheet(
    ws: Worksheet,
    reko_reports: list[RekoReport],
    incident_map: dict[uuid.UUID, Incident],
    personnel_map: dict[uuid.UUID, Personnel],
) -> None:
    """Add reko reports sheet."""
    headers = [
        "Report ID",
        "Einsatz ID",
        "Einsatz Titel",
        "Relevant",
        "Zusammenfassung",
        "Zusätzliche Notizen",
        "Stromversorgung",
        "Eingereicht Um",
        "Eingereicht Von",
        "Entwurf",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    for row_num, reko in enumerate(reko_reports, 2):
        incident = incident_map.get(reko.incident_id)
        personnel = personnel_map.get(reko.submitted_by_personnel_id) if reko.submitted_by_personnel_id else None

        ws.cell(row=row_num, column=1, value=str(reko.id))
        ws.cell(row=row_num, column=2, value=str(reko.incident_id))
        ws.cell(row=row_num, column=3, value=incident.title if incident else "")

        # is_relevant can be None (not yet determined), True, or False
        relevant_str = "" if reko.is_relevant is None else "Ja" if reko.is_relevant else "Nein"
        ws.cell(row=row_num, column=4, value=relevant_str)

        ws.cell(row=row_num, column=5, value=reko.summary_text or "")
        ws.cell(row=row_num, column=6, value=reko.additional_notes or "")
        ws.cell(row=row_num, column=7, value=reko.power_supply or "")
        ws.cell(row=row_num, column=8, value=format_timestamp(reko.submitted_at))
        ws.cell(row=row_num, column=9, value=personnel.name if personnel else "")
        ws.cell(row=row_num, column=10, value="Ja" if reko.is_draft else "Nein")


def get_safe_filename(event_name: str) -> str:
    """
    Create a safe filename from event name.

    Only ASCII alphanumeric characters are allowed in HTTP headers.
    Non-ASCII characters are replaced with underscores.
    """
    event_name_safe = "".join(c if c.isascii() and (c.isalnum() or c in (" ", "-", "_")) else "_" for c in event_name)
    # Collapse multiple underscores and strip trailing underscores
    event_name_safe = re.sub(r"_+", "_", event_name_safe).strip("_")
    return event_name_safe
