"""Excel import/export service for bulk data management."""

import asyncio
import logging
from collections.abc import Sequence
from dataclasses import dataclass, field
from datetime import UTC, datetime
from io import BytesIO
from typing import Any, Literal, TypedDict

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import Select, and_, delete, func, or_, select
from sqlalchemy import null as sa_null
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.sql.elements import ColumnElement

from ..models import (
    EventAttendance,
    EventSpecialFunction,
    Incident,
    IncidentAssignment,
    IncidentGroup,
    IncidentGroupAssignment,
    Material,
    Personnel,
    PersonnelExternalIdentity,
    Vehicle,
)
from .audit_export_service import EventReportData
from .pdf_report_service import (
    LOCAL_TZ,
    WorkWindow,
    board_personnel_count,
    extra_material_rows,
    format_extra_material,
    material_checklist_rows,
    material_left_on_site_names,
    material_used_label,
    rapport_by_incident,
    rapport_filing_lines,
    rapport_work_windows,
    vehicle_present_names,
)

# Column definitions
PERSONNEL_COLUMNS = [
    ("name", True),  # (column_name, required)
    ("role", False),
    ("status", False),
]

# The personnel column was called "availability" until the field was renamed to
# match Vehicle.status / Material.status. Workbooks exported before that are still
# accepted on import; only the header differs, the values never did.
LEGACY_PERSONNEL_COLUMNS = [("availability" if name == "status" else name) for name, _ in PERSONNEL_COLUMNS]

VEHICLE_COLUMNS = [
    ("name", True),
    ("type", True),
    ("display_order", True),
    ("status", True),
    ("radio_call_sign", True),
]

MATERIAL_COLUMNS = [
    ("name", True),
    ("type", True),
    ("location", True),
    ("description", False),
]

# Valid enum values. There is no VEHICLE_TYPES list any more: one sat here for years
# looking like an allowed-values check while enforcing nothing, and the docs had to tell
# operators that `type` is free text anyway. Free text is the intended behaviour – stations
# run "Anhänger", "Pikett-Bus" and whatever else is in the hall.
VEHICLE_STATUSES = ["available", "unavailable"]
PERSONNEL_STATUSES = ["available", "unavailable"]
# Material types are no longer hardcoded - validation now accepts any non-empty string

logger = logging.getLogger(__name__)


def _cell(value: object) -> str:
    """The operator's own cell value, quoted and capped, for an error message.

    Echoing it back is fine – it is their file – but a cell holds up to 32k
    characters and none of them belong in an HTTP response.
    """
    text = str(value)
    return f"'{text[:60]}…'" if len(text) > 60 else f"'{text}'"


class ExcelImportError(Exception):
    """A refused workbook, carrying the sheet and row the operator has to go fix.

    The parser has always known which cell it tripped over; the API used to throw
    that away and answer "Excel-Datei konnte nicht verarbeitet werden", which
    leaves a volunteer to bisect an 18-row spreadsheet by hand. So `message` is
    built to be returned verbatim: every `detail` below is a German literal
    authored in this module, interpolating only the uploaded file's own cells
    (via `_cell`) and our own column constants – never a path, a query or a
    traceback. The one raise site whose text was NOT ours – openpyxl's own
    exception string, wrapped as "Invalid Excel file: {e}" – logs the original
    and raises a fixed sentence instead, because nothing promises what a
    third-party parser puts in there.
    """

    def __init__(self, detail: str, *, sheet: str | None = None, row: int | None = None) -> None:
        self.detail = detail
        self.sheet = sheet
        self.row = row
        super().__init__(self.message)

    @property
    def message(self) -> str:
        """The whole complaint in one line, e.g. `Vehicles Zeile 7 – ungültiger Status 'x'`.

        Sheet names stay English: they name an actual tab in the operator's file.
        """
        if self.sheet is not None and self.row is not None:
            return f"{self.sheet} Zeile {self.row} – {self.detail}"
        if self.sheet is not None:
            return f"Blatt {self.sheet} – {self.detail}"
        return self.detail


@dataclass(frozen=True, slots=True)
class ParsedSheet:
    """One sheet's rows, plus whether that sheet was in the workbook at all.

    The distinction is the whole point: in `replace` mode an absent sheet means
    "this file says nothing about vehicles" while a present-but-empty sheet means
    "the station has no vehicles, clear the table". Both used to arrive here as an
    empty list, so a Personnel-only workbook deleted the fleet and the material
    inventory and reported `success: true`.
    """

    present: bool
    rows: list[dict[str, Any]] = field(default_factory=list)


@dataclass(frozen=True, slots=True)
class ParsedImport:
    """A validated workbook: one `ParsedSheet` per table the import can write."""

    personnel: ParsedSheet
    vehicles: ParsedSheet
    materials: ParsedSheet


def generate_empty_template() -> BytesIO:
    """Generate empty Excel template with example rows."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Personnel sheet
    ws_personnel = wb.create_sheet("Personnel")
    ws_personnel.append([col[0] for col in PERSONNEL_COLUMNS])
    # Header styling
    for cell in ws_personnel[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    # Example rows
    ws_personnel.append(["Max Mustermann", "Fahrer", "available"])
    ws_personnel.append(["Anna Schmidt", "", "unavailable"])

    # Vehicles sheet
    ws_vehicles = wb.create_sheet("Vehicles")
    ws_vehicles.append([col[0] for col in VEHICLE_COLUMNS])
    for cell in ws_vehicles[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws_vehicles.append(["TLF 1", "TLF", "1", "available", "Florian 1"])
    ws_vehicles.append(["DLK 1", "DLK", "2", "available", "Florian 2"])

    # Materials sheet
    ws_materials = wb.create_sheet("Materials")
    ws_materials.append([col[0] for col in MATERIAL_COLUMNS])
    for cell in ws_materials[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    # Example with duplicates showing multiple items
    ws_materials.append(["Tauchpumpe Gr.", "Tauchpumpen", "TLF", ""])
    ws_materials.append(["Tauchpumpe Kl.", "Tauchpumpen", "TLF", ""])
    ws_materials.append(["Wassersauger", "Wassersauger", "Pio", ""])

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _check_headers(sheet: str, headers: list[Any], expected: list[str]) -> None:
    """Refuse a sheet whose header row is not the one we know how to read."""
    if headers == expected:
        return
    # Only as many of the file's own headers as it takes to see the mismatch – a sheet
    # can carry a hundred stray columns and the error still has to fit in a toast.
    found = ", ".join(_cell(header) for header in headers[: len(expected) + 3]) or "keine"
    raise ExcelImportError(
        f"falsche Spaltenüberschriften. Erwartet: {', '.join(expected)}, gefunden: {found}.",
        sheet=sheet,
    )


def _parse_personnel_sheet(wb: Workbook) -> ParsedSheet:
    """Parse the Personnel sheet, or report it absent."""
    if "Personnel" not in wb.sheetnames:
        return ParsedSheet(present=False)

    ws = wb["Personnel"]
    headers = [cell.value for cell in ws[1]]
    expected_headers = [col[0] for col in PERSONNEL_COLUMNS]
    if headers == LEGACY_PERSONNEL_COLUMNS:
        headers = expected_headers
    _check_headers("Personnel", headers, expected_headers)

    rows: list[dict[str, Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue  # Skip empty rows

        row_data = dict(zip(expected_headers, row, strict=False))

        if not row_data.get("name"):
            raise ExcelImportError(
                "Spalte 'name' ist leer, jede Zeile braucht einen Namen.", sheet="Personnel", row=row_idx
            )

        if row_data.get("status") and row_data["status"] not in PERSONNEL_STATUSES:
            raise ExcelImportError(
                f"ungültiger Status {_cell(row_data['status'])}. Erlaubt: {', '.join(PERSONNEL_STATUSES)}.",
                sheet="Personnel",
                row=row_idx,
            )

        if not row_data.get("status"):
            row_data["status"] = "unavailable"

        rows.append(row_data)

    return ParsedSheet(present=True, rows=rows)


def _parse_vehicles_sheet(wb: Workbook) -> ParsedSheet:
    """Parse the Vehicles sheet, or report it absent.

    `type` is deliberately unvalidated: it is free text, and stations run
    "Anhänger", "Pikett-Bus" and other names no list here would have guessed.
    """
    if "Vehicles" not in wb.sheetnames:
        return ParsedSheet(present=False)

    ws = wb["Vehicles"]
    expected_headers = [col[0] for col in VEHICLE_COLUMNS]
    _check_headers("Vehicles", [cell.value for cell in ws[1]], expected_headers)

    rows: list[dict[str, Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue

        row_data = dict(zip(expected_headers, row, strict=False))

        for column, required in VEHICLE_COLUMNS:
            if required and not row_data.get(column):
                raise ExcelImportError(f"Spalte '{column}' ist leer (Pflichtfeld).", sheet="Vehicles", row=row_idx)

        if row_data["status"] not in VEHICLE_STATUSES:
            raise ExcelImportError(
                f"ungültiger Status {_cell(row_data['status'])}. Erlaubt: {', '.join(VEHICLE_STATUSES)}.",
                sheet="Vehicles",
                row=row_idx,
            )

        try:
            row_data["display_order"] = int(row_data["display_order"])
        except (ValueError, TypeError):
            raise ExcelImportError(
                f"'display_order' muss eine ganze Zahl sein, steht aber auf {_cell(row_data['display_order'])}.",
                sheet="Vehicles",
                row=row_idx,
            ) from None

        rows.append(row_data)

    return ParsedSheet(present=True, rows=rows)


def _parse_materials_sheet(wb: Workbook) -> ParsedSheet:
    """Parse the Materials sheet, or report it absent.

    `type` accepts any non-empty string – material categories differ per station.
    """
    if "Materials" not in wb.sheetnames:
        return ParsedSheet(present=False)

    ws = wb["Materials"]
    expected_headers = [col[0] for col in MATERIAL_COLUMNS]
    _check_headers("Materials", [cell.value for cell in ws[1]], expected_headers)

    rows: list[dict[str, Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue

        row_data = dict(zip(expected_headers, row, strict=False))

        for column, required in MATERIAL_COLUMNS:
            if required and not row_data.get(column):
                raise ExcelImportError(f"Spalte '{column}' ist leer (Pflichtfeld).", sheet="Materials", row=row_idx)

        rows.append(row_data)

    return ParsedSheet(present=True, rows=rows)


def validate_and_parse_excel(file_bytes: bytes) -> ParsedImport:
    """
    Validate and parse an uploaded workbook.

    A sheet the file does not contain comes back as `ParsedSheet(present=False)`,
    which is NOT the same as a sheet that is there and empty – see `ParsedSheet`.

    Raises ExcelImportError if validation fails; its `message` is safe to return
    to the caller and names the sheet and row.
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
    except Exception as e:
        # openpyxl's own text is not ours to hand out (it can name whatever the
        # library felt like naming), so it goes to the log and the operator gets
        # the one thing they can act on: the file is not a readable workbook.
        logger.warning("Rejected Excel upload openpyxl could not open: %s", e, exc_info=True)
        raise ExcelImportError(
            "Die Datei konnte nicht als Excel-Arbeitsmappe geöffnet werden. "
            "Ist es wirklich eine .xlsx-Datei (nicht CSV, nicht umbenannt)?"
        ) from e

    return ParsedImport(
        personnel=_parse_personnel_sheet(wb),
        vehicles=_parse_vehicles_sheet(wb),
        materials=_parse_materials_sheet(wb),
    )


class ImportDeletionImpact(TypedDict):
    """How many rows the chosen import mode would destroy, per entity.

    `incident_assignments` is the one nobody expects: assignments carry a bare
    polymorphic `resource_id` with no foreign key, so deleting a Personnel /
    Vehicle / Material row leaves the assignment behind pointing at nothing.
    `active_*` is the subset still on the board (`unassigned_at IS NULL`).

    `incident_group_assignments` is the SAME hazard one level up. Aufträge own
    their resources on `incident_group_assignments`, an identical polymorphic
    table, and it was missing here for exactly as long as it existed – a route
    with a full squad on it was invisible to both the preview and the refusal.

    The `cascade_*` numbers are the opposite failure: those rows do NOT survive,
    they vanish. `personnel.id` is a real foreign key with ON DELETE CASCADE in
    three tables, so wiping the roster takes every check-in of the running event
    with it, and the preview said nothing at all. Deleting a *vehicle* cascades
    into `event_special_functions` too (Fahrer), which is why that number is not
    a personnel-only count. Nothing references `materials.id`.
    """

    personnel: int
    vehicles: int
    materials: int
    incident_assignments: int
    active_incident_assignments: int
    incident_group_assignments: int
    active_incident_group_assignments: int
    cascade_event_attendance: int
    cascade_event_special_functions: int
    cascade_personnel_identities: int


# The two polymorphic assignment tables, which are byte-for-byte the same shape in
# every way this module cares about. Anything added here has to be added to
# `list_assignment_blockers` as well, or the refusal quotes a count it cannot break down.
_AssignmentTable = type[IncidentAssignment] | type[IncidentGroupAssignment]


def _assignments_to_live_resources(table: _AssignmentTable = IncidentAssignment) -> ColumnElement[bool]:
    """Assignment rows whose `resource_id` still resolves to an existing resource.

    A `replace` deletes every personnel, vehicle and material row, so every
    assignment matching this is an assignment that would be orphaned.
    """
    return or_(
        and_(
            table.resource_type == "personnel",
            table.resource_id.in_(select(Personnel.id)),
        ),
        and_(
            table.resource_type == "vehicle",
            table.resource_id.in_(select(Vehicle.id)),
        ),
        and_(
            table.resource_type == "material",
            table.resource_id.in_(select(Material.id)),
        ),
    )


async def count_deletion_impact(db: AsyncSession, mode: Literal["replace", "append"]) -> ImportDeletionImpact:
    """Count what `mode` would delete, so the preview can say it out loud.

    A station uploaded a one-row sheet to add two recruits and lost 18 people,
    5 vehicles and 26 materials – plus 24 assignments on three running incidents
    that then pointed at rows which no longer existed. The numbers were all
    knowable beforehand; nothing asked for them.

    Counted as a full three-table `replace`, the same way `personnel` / `vehicles`
    / `materials` already are, even though `import_data` only clears the tables
    whose sheet is present. `_refuse_missing_sheets` depends on exactly that: it
    needs to know a table has rows before it can refuse a workbook for omitting
    it. The numbers are therefore the worst case, and a workbook that omits a
    sheet never gets far enough to make them wrong.
    """
    if mode == "append":
        # append inserts and never deletes – there is nothing to warn about.
        return {
            "personnel": 0,
            "vehicles": 0,
            "materials": 0,
            "incident_assignments": 0,
            "active_incident_assignments": 0,
            "incident_group_assignments": 0,
            "active_incident_group_assignments": 0,
            "cascade_event_attendance": 0,
            "cascade_event_special_functions": 0,
            "cascade_personnel_identities": 0,
        }

    async def _count(stmt: Select[tuple[int]]) -> int:
        return (await db.execute(stmt)).scalar_one()

    async def _count_all(table: type[Any]) -> int:
        return await _count(select(func.count()).select_from(table))

    affected = _assignments_to_live_resources()
    affected_groups = _assignments_to_live_resources(IncidentGroupAssignment)
    return {
        "personnel": await _count_all(Personnel),
        "vehicles": await _count_all(Vehicle),
        "materials": await _count_all(Material),
        "incident_assignments": await _count(select(func.count()).select_from(IncidentAssignment).where(affected)),
        "active_incident_assignments": await _count(
            select(func.count())
            .select_from(IncidentAssignment)
            .where(affected, IncidentAssignment.unassigned_at.is_(None))
        ),
        "incident_group_assignments": await _count(
            select(func.count()).select_from(IncidentGroupAssignment).where(affected_groups)
        ),
        "active_incident_group_assignments": await _count(
            select(func.count())
            .select_from(IncidentGroupAssignment)
            .where(affected_groups, IncidentGroupAssignment.unassigned_at.is_(None))
        ),
        # Unconditional counts, not filtered by anything: `personnel_id` is NOT NULL
        # with ON DELETE CASCADE in all three, so emptying `personnel` empties them.
        "cascade_event_attendance": await _count_all(EventAttendance),
        "cascade_event_special_functions": await _count_all(EventSpecialFunction),
        "cascade_personnel_identities": await _count_all(PersonnelExternalIdentity),
    }


class AssignmentBlocker(TypedDict):
    """One incident or Auftrag still holding assignments that a `replace` would orphan.

    The count in `ImportDeletionImpact` is what the refusal used to say, and a
    number is useless to the person who has to act on it: they get "6 aktive
    Zuteilungen" and no way to find the six. This is the same set, broken down
    by incident and by resource type – enough to walk to the board and clear it.

    `materials` is separated out for a reason: "Alle freigeben" and completing an
    incident release personnel and vehicles only, so a material blocker survives
    exactly the action the operator reaches for first.

    `kind` exists because "release it on the incident" is the wrong instruction
    for half of these: an Auftrag owns its squad on the route, not on any of its
    stops, so an operator sent to the stop finds an empty resource list and a
    refusal that keeps coming back.
    """

    kind: Literal["incident", "group"]
    title: str
    location: str | None
    deleted: bool
    personnel: int
    vehicles: int
    materials: int
    total: int


def _fold_blocker_rows(
    rows: Sequence[Any],
    kind: Literal["incident", "group"],
) -> list[AssignmentBlocker]:
    """Fold `(id, title, location, deleted_at, resource_type, count)` rows into blockers."""
    folded: dict[Any, AssignmentBlocker] = {}
    for owner_id, title, location, deleted_at, resource_type, count in rows:
        blocker = folded.setdefault(
            owner_id,
            {
                "kind": kind,
                "title": title,
                "location": location,
                "deleted": deleted_at is not None,
                "personnel": 0,
                "vehicles": 0,
                "materials": 0,
                "total": 0,
            },
        )
        # The CHECK constraint on `resource_type` keeps this exhaustive.
        key = {"personnel": "personnel", "vehicle": "vehicles", "material": "materials"}[resource_type]
        blocker[key] += count  # type: ignore[literal-required]
        blocker["total"] += count
    return list(folded.values())


async def list_assignment_blockers(db: AsyncSession) -> list[AssignmentBlocker]:
    """Name what is behind the active-assignment counts, worst first.

    Counts the exact same rows as `count_deletion_impact` (active assignment,
    resource still exists) across BOTH assignment tables, so the totals here
    always add up to the number the refusal quotes – a list that is one short of
    the count sends the operator hunting for a blocker that does not exist.

    Soft-deleted incidents and Aufträge are in here on purpose. Deleting either
    does not release its resources, so the assignments keep blocking the import
    while the card is gone from the board; `deleted` lets the message say so
    instead of naming something nobody can find.
    """
    incident_rows = (
        await db.execute(
            select(
                Incident.id,
                Incident.title,
                Incident.location_address,
                Incident.deleted_at,
                IncidentAssignment.resource_type,
                func.count().label("n"),
            )
            .select_from(IncidentAssignment)
            # Inner join is safe: `incident_assignments.incident_id` is a real FK
            # with ON DELETE CASCADE – unlike `resource_id`, which is the bare UUID
            # this whole refusal exists because of.
            .join(Incident, Incident.id == IncidentAssignment.incident_id)
            .where(_assignments_to_live_resources(), IncidentAssignment.unassigned_at.is_(None))
            .group_by(
                Incident.id,
                Incident.title,
                Incident.location_address,
                Incident.deleted_at,
                IncidentAssignment.resource_type,
            )
        )
    ).all()

    # An Auftrag has no address of its own – its stops carry the addresses – so the
    # location column is a literal NULL rather than a join we would have to aggregate.
    group_rows = (
        await db.execute(
            select(
                IncidentGroup.id,
                IncidentGroup.name,
                sa_null(),
                IncidentGroup.deleted_at,
                IncidentGroupAssignment.resource_type,
                func.count().label("n"),
            )
            .select_from(IncidentGroupAssignment)
            .join(IncidentGroup, IncidentGroup.id == IncidentGroupAssignment.incident_group_id)
            .where(
                _assignments_to_live_resources(IncidentGroupAssignment),
                IncidentGroupAssignment.unassigned_at.is_(None),
            )
            .group_by(
                IncidentGroup.id,
                IncidentGroup.name,
                IncidentGroup.deleted_at,
                IncidentGroupAssignment.resource_type,
            )
        )
    ).all()

    blockers = _fold_blocker_rows(incident_rows, "incident") + _fold_blocker_rows(group_rows, "group")

    # Worst first, so the capped list names what is worth walking to; kind and title
    # as tie-breaks only so the order is stable between two identical calls.
    return sorted(blockers, key=lambda b: (-b["total"], b["kind"], b["title"]))


async def import_data(
    db: AsyncSession,
    parsed_data: ParsedImport,
    mode: Literal["replace", "append"],
    user_id: str,
) -> dict[str, int]:
    """
    Import parsed data into database.

    Modes:
    - replace: Delete the tables the workbook actually contains, insert new
    - append: Keep existing, add new

    There used to be a third mode, `merge`, documented as "update existing by
    name, add new". It never did that – it ran the same three DELETEs as
    `replace`. The endpoint rejects it now rather than keep a name that means
    the opposite of what it does.

    Returns counts: {personnel: X, vehicles: Y, materials: Z}
    """
    counts = {"personnel": 0, "vehicles": 0, "materials": 0}

    if mode == "replace":
        # Only the tables whose sheet is in the file. `replace` used to DELETE all
        # three unconditionally, so a Personnel-only workbook wiped the fleet and the
        # material inventory and put nothing back – reported as success. A sheet that
        # is present and empty still clears its table: that is how a station empties
        # one on purpose. The endpoint refuses the absent-sheet case outright rather
        # than silently leaving a table behind (see `admin.execute_excel_import`);
        # this branch is the second lock on the same door.
        if parsed_data.personnel.present:
            await db.execute(delete(Personnel))
        if parsed_data.vehicles.present:
            await db.execute(delete(Vehicle))
        if parsed_data.materials.present:
            await db.execute(delete(Material))
        await db.commit()

    # Insert personnel
    for person_data in parsed_data.personnel.rows:
        personnel = Personnel(**person_data)
        db.add(personnel)
        counts["personnel"] += 1

    # Insert vehicles
    for vehicle_data in parsed_data.vehicles.rows:
        vehicle = Vehicle(**vehicle_data)
        db.add(vehicle)
        counts["vehicles"] += 1

    # Insert materials (duplicate rows = multiple items)
    for material_data in parsed_data.materials.rows:
        material = Material(**material_data)
        db.add(material)
        counts["materials"] += 1

    await db.commit()

    return counts


async def export_data_to_excel(db: AsyncSession) -> BytesIO:
    """Export all personnel, vehicles, and materials to Excel."""
    personnel_result = await db.execute(select(Personnel).order_by(Personnel.name))
    personnel = personnel_result.scalars().all()
    vehicle_result = await db.execute(select(Vehicle).order_by(Vehicle.display_order))
    vehicles = vehicle_result.scalars().all()
    material_result = await db.execute(select(Material).order_by(Material.location, Material.name))
    materials = material_result.scalars().all()

    # openpyxl workbook building is pure CPU — keep it off the event loop
    # so a large export doesn't freeze every operator's requests (audit H4).
    return await asyncio.to_thread(_build_export_workbook, personnel, vehicles, materials)


def _build_export_workbook(
    personnel: Sequence[Personnel], vehicles: Sequence[Vehicle], materials: Sequence[Material]
) -> BytesIO:
    """Blocking workbook construction — runs in a worker thread."""
    wb = Workbook()
    wb.remove(wb.active)

    # Personnel sheet
    ws_personnel = wb.create_sheet("Personnel")
    ws_personnel.append([col[0] for col in PERSONNEL_COLUMNS])
    for cell in ws_personnel[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for person in personnel:
        ws_personnel.append(
            [
                person.name,
                person.role or "",
                person.status,
            ]
        )

    # Vehicles sheet
    ws_vehicles = wb.create_sheet("Vehicles")
    ws_vehicles.append([col[0] for col in VEHICLE_COLUMNS])
    for cell in ws_vehicles[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for vehicle in vehicles:
        ws_vehicles.append(
            [
                vehicle.name,
                vehicle.type,
                vehicle.display_order,
                vehicle.status,
                vehicle.radio_call_sign,
            ]
        )

    # Materials sheet
    ws_materials = wb.create_sheet("Materials")
    ws_materials.append([col[0] for col in MATERIAL_COLUMNS])
    for cell in ws_materials[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for material in materials:
        ws_materials.append(
            [
                material.name,
                material.type,
                material.location,
                material.description or "",
            ]
        )

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# Einsätze export (plan 25, §7 / decision 21)
#
# One wide row per Schadenplatz. It matches **no** external format on purpose:
# somebody retypes this into the billing system by hand, and that system is
# archaic enough that mirroring its layout buys nothing — so the sheet optimises
# for being *read while retyping*: human headers, everything about one
# Schadenplatz on one line, and no derived person-hours column (nothing
# downstream asked for one; `cost_snapshot_json` keeps the per-person from/to if
# that ever changes). It just does not need that name on it: the crew filling
# the slip is recording an Einsatz, not writing an invoice.
# ---------------------------------------------------------------------------

# (header, column width), in the reading order of the paper slip.
EINSAETZE_COLUMNS: list[tuple[str, int]] = [
    ("Einsatz-Nr.", 11),
    ("Adresse", 34),
    ("Beginn", 17),
    ("Ende", 17),
    ("Dauer", 9),
    ("Personal", 9),
    ("Personal korrigiert", 20),
    ("Fahrzeuge", 34),
    # Two columns since §18.31, matching the two inputs the form asks for. The
    # phone is its own column for the same reason it is its own field: whoever
    # writes the invoices sorts and dials it, and it cannot do either from
    # inside a paragraph.
    ("Eigentümer / Halter", 46),
    ("Eigentümer / Halter Telefon", 22),
    ("Material gebraucht", 46),
    ("Material vor Ort verblieben", 34),
    ("Weiteres Material", 28),
    ("Kurzbericht", 60),
    ("Erfasst von", 44),
]


def _local_dt(value: datetime | None) -> str:
    """``DD.MM.YYYY HH:MM`` in Swiss local time, or empty.

    This sheet is read by a person with a wall clock, not by a machine — hence
    local time here where the audit export uses ISO 8601.
    """
    if value is None:
        return ""
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.astimezone(LOCAL_TZ).strftime("%d.%m.%Y %H:%M")


def format_duration(start: datetime | None, end: datetime | None) -> str:
    """``h:mm`` between the two Tätigkeit timestamps, or empty.

    Computed, never stored, so it always agrees with the Beginn/Ende on its own
    row — which are themselves derived from the board (`rapport_work_windows`).
    """
    if start is None or end is None:
        return ""
    if start.tzinfo is None:
        start = start.replace(tzinfo=UTC)
    if end.tzinfo is None:
        end = end.replace(tzinfo=UTC)
    minutes = int((end - start).total_seconds() // 60)
    if minutes < 0:
        return ""
    return f"{minutes // 60}:{minutes % 60:02d}"


def _corrected_cell(corrected: bool, board_value: int) -> str:
    """``Ja (Board: 6)`` — the flag plus the number the crew disagreed with.

    The divergence is itself information (decision 5): it says the board was
    behind reality, and a corrected count without the board's own number next
    to it is just a number.
    """
    return f"Ja (Board: {board_value})" if corrected else ""


def build_einsaetze_workbook(data: EventReportData) -> BytesIO:
    """One row per Schadenplatz, including the ones **without** a rapport.

    A missing rapport is a blank row carrying its address, never a missing row:
    there is no acceptance step by design (decision 10), so the gaps have to be
    visible to whoever writes the invoices.
    """
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Einsätze")

    for col_num, (header, width) in enumerate(EINSAETZE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="top")
        ws.column_dimensions[cell.column_letter].width = width
    ws.freeze_panes = "A2"

    reports = rapport_by_incident(data)
    # Beginn/Ende are derived from the board once for the whole sheet — the crew
    # no longer types them, and this walks the already-loaded assignments and
    # transitions rather than querying per row.
    windows = rapport_work_windows(data)

    for row_num, (index, incident) in enumerate(enumerate(data.incidents, 1), 2):
        report = reports.get(incident.id)

        values: list[Any]
        if report is None:
            values = [index, incident.location_address or "", *[""] * (len(EINSAETZE_COLUMNS) - 2)]
        else:
            window = windows.get(incident.id, WorkWindow(None, None))
            values = [
                index,
                incident.location_address or "",
                _local_dt(window.started_at),
                _local_dt(window.ended_at),
                format_duration(window.started_at, window.ended_at),
                report.personnel_count if report.personnel_count is not None else "",
                _corrected_cell(report.personnel_count_corrected, board_personnel_count(data, incident.id)),
                # The vehicles the crew ticked, by name: which ones were there is
                # the question the billing side actually asks, and a number
                # answers it for nobody.
                ", ".join(vehicle_present_names(report)),
                report.owner_name or "",
                report.owner_phone or "",
                # Every unit with its own answer, "nicht gebraucht" included
                # (decision 16). There is no third state any more (§18.32): the
                # tick is prefilled ja and the crew unticks the exceptions.
                # Consumables carry no left-on-site state at all, which is why
                # that lives in its own column (decision 26).
                "; ".join(
                    f"{row.get('name') or '?'}: {material_used_label(row.get('used'))}"
                    for row in material_checklist_rows(report)
                ),
                ", ".join(material_left_on_site_names(report)),
                # Every "Weiteres Material" entry with its own on-site answer
                # (§18.35). It stays in its own column rather than joining the
                # one left of it: these are names the crew wrote, not units the
                # board dispatched, and a billing reader must be able to tell
                # the two apart at a glance.
                ", ".join(format_extra_material(row) for row in extra_material_rows(report)),
                report.kurzbericht or "",
                " · ".join(rapport_filing_lines(data, report)),
            ]

        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def export_einsaetze_excel(data: EventReportData) -> BytesIO:
    """Build the Einsätze workbook off the event loop (audit H4)."""
    return await asyncio.to_thread(build_einsaetze_workbook, data)
